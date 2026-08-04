"""
Courtage-Extraktor fuer SSH Versicherungsmakler GmbH & Co. KG
================================================================

Liest alle Courtageabrechnungs-PDFs eines Monatsordners (z.B. "Juni-2026")
und extrahiert je Buchungszeile: Versicherer, Kunde, Provision.

Aufruf:
    python courtage_extraktor.py Juni-2026

Ergebnis:
    Courtage-Tool/output/Juni-2026/Kunde_Provision_Juni-2026.xlsx

Die Zuordnung Kunde -> Betreuer (Courtage-Verteilung) ist bewusst NICHT
Teil dieses Skripts (das ist der naechste Ausbauschritt). Dieses Skript
liefert nur die Rohdaten je Buchung: Kunde + Provision + Herkunft.

Aufbau, kurz:
- Manche Versicherer liefern eine echte Positionstabelle (Kunde + Betrag
  pro Zeile) -> generische, koordinatenbasierte Tabellen-Engine.
- Manche Versicherer (aktuell: Fondsfinanz) liefern ein Block-Format
  (ein Kundenblock ueber mehrere Zeilen) -> eigener Parser.
- Manche PDFs enthalten keinen Text (Scan) -> OCR-Fallback (Tesseract).
- Manche PDFs sind reine Sammelabrechnungen ohne Kundenbezug (z.B. ARAG,
  VEMA-Pool-Uebersicht) -> werden als "Sammelbeleg" gekennzeichnet, kein
  Kunde wird erfunden.
- Manche PDFs (Allianz, AXA, Gothaer) haben einen Sonderfall in der
  PDF-Erzeugung (Text pro Seite komplett gespiegelt/verdreht) -> eigene,
  koordinatenbasierte Parser (siehe extract_allianz/extract_axa/
  extract_gothaer).
- Dialog und Continentale liefern echte Kundenpositionstabellen, aber im
  Scan zu klein/dicht/verschachtelt fuer zuverlaessige OCR -> bleiben
  manuelle Pruefung statt falsche Zahlen zu riskieren (bei Dialog zusaetzlich
  ein OCR-Kontrollbetrag, siehe find_dialog_total_hint).
- VEMA-Pool liefert Kundenpositionen nicht im PDF, sondern separat als CSV
  (VEMA-Poolabrechnung-N.csv) -> wird automatisch mitverarbeitet, wenn sie
  im Monatsordner liegt bzw. mit hochgeladen wird (siehe extract_vema_csv).
"""

import csv
import glob
import os
import re
import sys
from collections import Counter

import pdfplumber
import pandas as pd

try:
    import pytesseract
    from PIL import ImageOps
    TESSERACT_EXE = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(TESSERACT_EXE):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE
    TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
    LOCAL_TESSDATA = os.path.join(TOOL_DIR, "tessdata")
    if os.path.isdir(LOCAL_TESSDATA) and os.listdir(LOCAL_TESSDATA):
        # Lokales Sprachpaket vorhanden (Windows-Installation ohne Schreib-
        # zugriff auf das Tesseract-Programmverzeichnis) - andernfalls nutzt
        # Tesseract sein eigenes System-Tessdata-Verzeichnis, z.B. auf
        # Streamlit Cloud (Linux) ueber das apt-Paket tesseract-ocr-deu
        # (siehe packages.txt).
        os.environ["TESSDATA_PREFIX"] = LOCAL_TESSDATA
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # .../Umsatz
TOOL_DIR = os.path.dirname(os.path.abspath(__file__))                   # .../Umsatz/Courtage-Tool

# ---------------------------------------------------------------------------
# Konfiguration: Schluesselwoerter fuer die generische Tabellen-Engine
# ---------------------------------------------------------------------------

# Tier 1: eindeutige, volle Schluesselwoerter - werden immer zuerst versucht.
NAME_KEYWORDS_STRONG = [
    "versicherungsnehmer", "kunde", "name, vorname", "vertragsinhaber",
    "versicherter", "name vn", "mandant",
]
# Tier 2: nur als Fallback, wenn Tier 1 nichts findet (z.B. Kopfzeile ueber
# zwei Zeilen gebrochen: "Versicherungs-" / "nehmer"). "versicherungs-"
# kollidiert sonst mit einer Spalte "Versicherungs-Nr.". "vn/tu" ist die in
# manchen Abrechnungen (z.B. HDI-Leben: "VN/TU") uebliche Abkuerzung fuer
# Versicherungsnehmer - bewusst NICHT das kuerzere "vn" allein, das z.B. in
# Kontaktangaben wie "Prov./Court.(VN):" faelschlich matchen wuerde.
# "vertragsinformationen" (Mannheimer) ist keine Namens-Spaltenueberschrift
# im engeren Sinn, aber die Spalte, in der der Name/die Firma tatsaechlich
# steht.
NAME_KEYWORDS_WEAK = ["versicherungs-", "name", "vn/tu", "vertragsinformationen"]

# Woerter, die in manchen Abrechnungen (z.B. Mannheimer) an der Position der
# Namens-Spalte auftauchen, aber Buchungsart-Label sind, kein Kundenname
# (z.B. eine eigene Zeile "Bestandspflege" oder "Folgebeitrag" zwischen der
# echten Namenszeile und der Betragszeile). Wird zusaetzlich zur
# Alphalauf-Pruefung angewendet, da diese Woerter lang genug sind, um die
# Alphalauf-Pruefung fuer sich allein zu bestehen.
NON_NAME_WORDS = {
    "bestandspflege", "folgebeitrag", "erstbeitrag", "nachtrag",
    "vers.schein", "praemienregul.haftpflicht", "beitrag/gebuehr/zahlung",
    "fremdversicherungsnummer",
    # ohne fuehrenden Buchstaben, da manche PDFs "ae" als kaputtes "�"
    # rendern (siehe UMLAUT_MAP-Hinweis) - Teilstring-Vergleich statt exakt.
    "lter90tage", "provisioni",
}

AMOUNT_KEYWORDS = [
    # Hinweis: "verguet" statt "vergüt", weil normalize() Umlaute VOR dem
    # Vergleich in ue/oe/ae/ss uebersetzt (siehe UMLAUT_MAP/normalize()).
    "courtage", "provision", "verguet", "betrag", "gesamtverg",
    # "buch-" statt "buchwert", weil das Wort manchmal ueber zwei Zeilen
    # getrennt ist ("Buch-" / "wert"). Mit Bindestrich, damit es nicht mit
    # "Buchungs-" (Datum/Typ-Spalte, keine Betragsspalte) kollidiert.
    "buch-",
    # "prov." (mit Punkt, z.B. Itzehoer: "Prov.(€)") statt "provision" -
    # manche Abrechnungen kuerzen die Spaltenueberschrift ab. "%" wird ueber
    # AMOUNT_EXCLUDE separat ausgeschlossen (z.B. "Prov.(%)").
    "prov.",
    # "gutschrift" (z.B. Alte Leipziger: "Gutschrift/Belastung").
    "gutschrift",
]
AMOUNT_EXCLUDE = ["%", "satz", "-satz", "erwart"]

BLACKLIST_LINE_KEYWORDS = [
    "summe", "saldo", "gesamt", "endsaldo", "kontostand", "uebertrag",
    "vortrag", "zwischensumme", "kontenstand",
    "auszahlungsbetrag", "stornoreserve", "abrechnungsbetrag",
    "seite", "kontoauszug", "davon", "total", "bezeichnung",
    "jahreswerte", "zahlungsausgang",
]
# Nur DIESE (Teilmenge der Blacklist) setzen den "aktuellen Kunden" auch
# zurueck, statt die Zeile nur zu ueberspringen: sie markieren ein echtes
# Abschnittsende (z.B. eine Spaltenkopf-Wiederholung wie "VGART Bezeichnung
# ..." bei Deurag, nach der die naechste Zeile sonst faelschlich dem letzten
# Kunden zugerechnet wuerde). Andere Blacklist-Woerter wie "uebertrag" oder
# "kontostand" sind dagegen oft nur eine harmlose Salden-Notiz MITTEN in
# einem mehrzeiligen Kundenblock (z.B. HDI: "Uebertrag 6.941,03 738,23"
# gefolgt von einer weiteren Buchungszeile DESSELBEN Kunden) - dort wuerde
# ein Reset eine legitime Fortsetzungszeile verwerfen.
RESET_CUSTOMER_KEYWORDS = ["bezeichnung", "davon", "jahreswerte"]
# Ausnahmen: diese Zeilen ENTHALTEN "summe", zaehlen aber trotzdem als
# gueltige Kundenzeile (Fondsfinanz-Blockformat: "Vertragssumme EUR X")
BLACKLIST_EXCEPTIONS = ["vertragssumme"]

# Ein einzelner Buchungsbetrag ueber dieser Schwelle ist bei diesem Makler-
# volumen unplausibel (typische Courtage-Betraege liegen im ein- bis
# niedrigen vierstelligen Bereich) - eher ein Fehltreffer (z.B. eine
# Handelsregister- oder Kontonummer, die zufaellig in der Betrags-Spalte
# liegt). Solche Zeilen werden verworfen statt falsche Zahlen zu liefern.
MAX_PLAUSIBLE_AMOUNT = 20000

TOTAL_LINE_PATTERNS = [
    r"Endsumme[n]?.{0,40}?([\d.,]+)\s*$",
    r"Gesamtbetrag[:]?\s*([\d.,]+)",
    r"Auszahlungsbetrag[:]?\s*([\d.,]+)",
    r"Abrechnungsbetrag[:]?\s*([\d.,]+)",
    r"Gesamt-Auszahlungsbetrag[:]?\s*([\d.,]+)",
    r"Summe Vermittler.*?([\-\d.,]+)\s*$",
    # Erste Zahl nach "Auszahlung" (nicht letzte!): manche Abrechnungen
    # zeigen "aktueller Zeitraum | kumuliert" nebeneinander - die kumulierte
    # Spalte (meist die letzte Zahl der Zeile) waere sonst falsch.
    r"Auszahlung\s+(?:\d\d\.\d\d\.\d{4}\s+)?([\-\d.,]+)",
    r"Wir überweisen[^0-9]*([\d.,]+)",
    r"Saldo (?:neu|zum \d\d\.\d\d\.\d{4})[:]?\s*([\-\d.,]+)",
]

REVERSED_MARKERS = [
    "gnunhcerbA", "ztasmU", "gnurehcisreV", "esiewnihztasuZ",
    "remhensgnurehcisreV", "gnunneK",
]

MONTHS_DE = [
    "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember",
]


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def parse_amount(text):
    """Parst deutsche (1.234,56 / 12,34-) und englische (1,234.56 / -12.34)
    Zahlenformate zu float. Gibt None zurueck, wenn kein plausibler Betrag."""
    if not text:
        return None
    s = text.strip()
    s = s.replace("€", "").replace("EUR", "").replace("�", "").strip()
    if not s:
        return None

    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:]
    if s.endswith("-"):
        neg = True
        s = s[:-1]
    s = s.strip()
    if not re.fullmatch(r"[\d.,]+", s):
        return None

    last_comma = s.rfind(",")
    last_dot = s.rfind(".")
    if last_comma == -1 and last_dot == -1:
        if not s.isdigit():
            return None
        val = float(s)
    elif last_comma > last_dot:
        # deutsches Format: Punkt = Tausender, Komma = Dezimal
        s2 = s.replace(".", "").replace(",", ".")
        try:
            val = float(s2)
        except ValueError:
            return None
    elif last_dot > last_comma:
        # englisches Format: Komma = Tausender, Punkt = Dezimal
        s2 = s.replace(",", "")
        try:
            val = float(s2)
        except ValueError:
            return None
    else:
        try:
            val = float(s)
        except ValueError:
            return None

    return -val if neg else val


UMLAUT_MAP = str.maketrans({
    "ü": "ue", "ö": "oe", "ä": "ae", "ß": "ss",
    "Ü": "ue", "Ö": "oe", "Ä": "ae",
})


def normalize(word):
    return word.lower().strip().translate(UMLAUT_MAP)


def matches_any(word_lower, keywords):
    return any(k in word_lower for k in keywords)


def _blacklist_check(text):
    if any(exc in text for exc in BLACKLIST_EXCEPTIONS):
        return False
    return any(k in text for k in BLACKLIST_LINE_KEYWORDS)


def is_blacklisted_line(line_text_lower):
    """Prueft eine Zeile gegen die Blacklist. Prueft zusaetzlich eine um
    Ziffern bereinigte Variante, da manche PDFs Kopf-/Fusszeilen-Texte
    (z.B. eine Positionsnummer) mit Wort-Ueberlagerung zeichenweise
    ineinander schieben, z.B. "3S0u1m2m11e" statt "Summe" - sowie eine
    Leerzeichen-bereinigte Variante, da manche PDFs Woerter mit
    Buchstaben-Spatium rendern, z.B. "S U M M E" statt "Summe"."""
    normalized = line_text_lower.translate(UMLAUT_MAP)
    if _blacklist_check(normalized):
        return True
    if _blacklist_check(re.sub(r"\d", "", normalized)):
        return True
    return _blacklist_check(normalized.replace(" ", ""))


def cluster_lines(words, tol=2.5):
    """Gruppiert Woerter nach vertikaler Position (top) zu Zeilen."""
    lines = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        placed = False
        for line in lines:
            if abs(line["top"] - w["top"]) <= tol:
                line["words"].append(w)
                line["top"] = (line["top"] * len(line["words"]) + w["top"]) / (len(line["words"]) + 1)
                placed = True
                break
        if not placed:
            lines.append({"top": w["top"], "words": [w]})
    for line in lines:
        line["words"].sort(key=lambda w: w["x0"])
    lines.sort(key=lambda l: l["top"])
    return lines


def get_words_for_page(page, resolution=300):
    """Liefert Woerter (text,x0,x1,top,bottom) einer Seite - aus der Text-
    ebene, oder per OCR falls die Seite keinen Text enthaelt (Scan)."""
    if len(page.chars) > 0:
        words = page.extract_words(x_tolerance=1.5, keep_blank_chars=False)
        return words, "text"

    if not OCR_AVAILABLE:
        return [], "kein-ocr"

    scale = 72.0 / resolution
    im = page.to_image(resolution=resolution).original
    data = pytesseract.image_to_data(im, lang="deu", output_type=pytesseract.Output.DICT)
    words = []
    n = len(data["text"])
    for i in range(n):
        txt = data["text"][i].strip()
        if not txt:
            continue
        conf = data.get("conf", ["-1"] * n)[i]
        try:
            if float(conf) < 25:
                continue
        except (ValueError, TypeError):
            pass
        x, y, w_, h_ = data["left"][i], data["top"][i], data["width"][i], data["height"][i]
        words.append({
            "text": txt,
            "x0": x * scale,
            "x1": (x + w_) * scale,
            "top": y * scale,
            "bottom": (y + h_) * scale,
        })
    return words, "ocr"


def _find_name_and_amount(combined_words):
    """Liefert (name_word, amount_word) oder None, falls in den gegebenen
    Woertern nicht beides eindeutig vorkommt. Nutzt zuerst die eindeutigen
    (Tier-1) Namens-Schluesselwoerter, erst danach die mehrdeutigeren
    Tier-2-Woerter (die z.B. mit 'Versicherungs-Nr.' kollidieren koennen)."""
    amount_candidates = [
        w for w in combined_words
        if matches_any(normalize(w["text"]), AMOUNT_KEYWORDS)
        and not matches_any(normalize(w["text"]), AMOUNT_EXCLUDE)
    ]
    if not amount_candidates:
        return None

    for name_keywords in (NAME_KEYWORDS_STRONG, NAME_KEYWORDS_WEAK):
        name_candidates = [w for w in combined_words if matches_any(normalize(w["text"]), name_keywords)]
        if name_candidates:
            name_word = min(name_candidates, key=lambda w: w["x0"])
            amount_word = max(amount_candidates, key=lambda w: w["x0"])
            return name_word, amount_word
    return None


def _header_from_span(lines, start, span):
    end = min(start + span, len(lines))
    combined = []
    for k in range(start, end):
        for w in lines[k]["words"]:
            tagged = dict(w)
            tagged["_line"] = k
            combined.append(tagged)
    if not combined:
        return None
    found = _find_name_and_amount(combined)
    if not found:
        return None
    name_word, amount_word = found

    def right_boundary(word, all_words):
        # nur Woerter aus derselben Zeile wie das gefundene Schluesselwort
        # heranziehen, damit fachfremde Woerter aus einer bei groesseren
        # Fenstern zufaellig mit erfassten Nachbarzeile die Spaltenbreite
        # nicht verfaelschen.
        same_line = [w for w in all_words if w["_line"] == word["_line"]]
        candidates = [w["x0"] for w in same_line if w["x0"] > word["x0"] + 3]
        # kleiner Sicherheitsabstand: Kopfzeilen-Wort und Datenzeilen-Wert
        # sind selten pixelgenau gleich ausgerichtet - ohne Puffer wuerde ein
        # Datenwert, der zufaellig 0,1pt vor der naechsten Spaltenueberschrift
        # beginnt, faelschlich noch in diese Spalte hineingezaehlt.
        return (min(candidates) - 2) if candidates else word["x1"] + 250

    def left_boundary(word, all_words):
        # Zahlen-Spalten sind meist rechtsbuendig: ein groesserer Betrag
        # (mehr Ziffern) beginnt weiter links als die Kopfzeilen-Ueberschrift
        # selbst. Als linke Grenze daher das rechte Ende der VORHERGEHENDEN
        # Spalte in derselben Zeile nehmen (mit etwas Puffer), statt der
        # eigenen x0 der Ueberschrift - sonst werden breite Zahlen
        # abgeschnitten.
        same_line = [w for w in all_words if w["_line"] == word["_line"]]
        candidates = [w["x1"] for w in same_line if w["x1"] < word["x0"] - 3]
        return max(candidates) + 2 if candidates else max(0, word["x0"] - 5)

    # Nur fuer die Betragsspalte auf die vorherige Spalte stuetzen (rechts-
    # buendige, unterschiedlich breite Zahlen wie "1.134,16" wuerden sonst
    # abgeschnitten). Fuer die Namensspalte NICHT, da die Namens-/Produkt-
    # Spalte oft eine kurze Kopfzeilen-Ueberschrift (z.B. "VSNR") aber viel
    # laengere Dateninhalte (z.B. lange Produktnamen) hat - dort wuerde die
    # vorige-Spalte-Grenze faelschlich in die Nachbarspalte hineinragen
    # (siehe R+V: "VSNR" ist kurz, "Produktname" in derselben Spalte lang).
    name_left = max(0, name_word["x0"] - 5)
    name_right = right_boundary(name_word, combined)
    amount_left = left_boundary(amount_word, combined)
    amount_right = right_boundary(amount_word, combined)
    return {
        "name_range": (name_left, name_right),
        "amount_range": (amount_left, amount_right),
    }, end


def extract_generic_table(pages_words):
    """Generische koordinatenbasierte Extraktion fuer 'normale' Tabellen-PDFs.
    pages_words: Liste von (page_index, words, source) Tupeln.

    Kopfzeilen werden nicht nur einmal pro Seite gesucht, sondern bei jeder
    Zeile neu geprueft: manche Versicherer (z.B. Amex-Pool) haben mehrere
    Teiltabellen mit unterschiedlichen Spaltenbreiten auf derselben Seite
    (z.B. 'Abschlussprovisionen' und 'Folgeprovisionen')."""
    rows = []
    columns = None
    current_customer = None

    for page_idx, words, source in pages_words:
        if not words:
            continue
        lines = cluster_lines(words)
        li = 0
        n = len(lines)

        while li < n:
            header_found = None
            for span in (1, 2, 3):
                result = _header_from_span(lines, li, span)
                if result:
                    header_found = result
                    break
            if header_found:
                columns, next_li = header_found
                li = next_li
                continue

            if columns is None:
                li += 1
                continue

            line = lines[li]
            line_text = " ".join(w["text"] for w in line["words"])
            line_lower = line_text.lower()
            if is_blacklisted_line(line_lower):
                if any(k in line_lower for k in RESET_CUSTOMER_KEYWORDS):
                    # Echtes Abschnittsende (siehe RESET_CUSTOMER_KEYWORDS):
                    # aktuellen Kunden zuruecksetzen, sonst wuerde eine
                    # nachfolgende, selbst nicht geblacklistete Rekap-Zeile
                    # (z.B. eine Aufschluesselung nach Vertragsart/Jahres-
                    # werten, die denselben Betrag nochmal zusammenfasst)
                    # faelschlich per Fortschreibung dem letzten echten
                    # Kunden zugerechnet und so doppelt gezaehlt.
                    current_customer = None
                li += 1
                continue

            name_l, name_r = columns["name_range"]
            amt_l, amt_r = columns["amount_range"]
            name_words = [w for w in line["words"] if name_l <= w["x0"] < name_r]
            amount_words = [w for w in line["words"] if amt_l <= w["x0"] < amt_r]
            other_words = [
                w for w in line["words"]
                if w not in name_words and w not in amount_words
            ]

            name_text = " ".join(w["text"] for w in name_words).strip(" ,")
            if name_text:
                letter_runs = re.findall(r"[A-Za-zÀ-ÿ]+", name_text)
                longest_run = max((len(r) for r in letter_runs), default=0)
                name_joined = normalize(name_text).replace(" ", "")
                if longest_run < 3:
                    # kein echter Kundenname: z.B. eine verirrte Saldozahl,
                    # oder ein Fragment einer Folgezeile (Datum/Kennzeichen
                    # wie "4 01.07.26 184,76 FB"), das zufaellig in die
                    # Namensspalte faellt.
                    name_text = ""
                elif any(nn in name_joined for nn in NON_NAME_WORDS):
                    # Buchungsart-/Spaltenkopf-Label (z.B. "Bestandspflege",
                    # "Erstbeitrag Vers.Schein"), kein Kundenname - siehe
                    # NON_NAME_WORDS. Teilstring-Vergleich (nicht exakte
                    # Gleichheit), da diese Woerter manchmal an einen
                    # zusaetzlichen echten Namensteil angrenzen.
                    name_text = ""
                elif re.match(r"^[A-ZÀ-Ü]{2,4}-[A-ZÀ-Ü]{1,4}(\s+\d+)?$", name_text):
                    # Kurzes Risikoort-/Produkt-Kuerzel mit Bindestrich (z.B.
                    # "HAL-SL 321", "LU-SW 312"), kein Kundenname - manche
                    # Abrechnungen (z.B. HDI) verwenden dieselbe Spalten-
                    # position auf Folgezeilen fuer ein anderes Feld.
                    name_text = ""

            # Von rechts nach links das erste als Zahl parsebare Wort in der
            # Betragsspalte nehmen - nicht nur die letzten ein/zwei Woerter
            # pruefen: manche Abrechnungen (z.B. Blaudirekt) haengen nach dem
            # Betrag noch einen Hinweis wie "Festbetrag" an, der sonst die
            # Zeile faelschlich verwerfen wuerde (Betrag stuende dann drei
            # statt zwei Woerter vor Zeilenende).
            amount_val = None
            for w in reversed(amount_words):
                amount_val = parse_amount(w["text"])
                if amount_val is not None:
                    break

            if name_text:
                current_customer = name_text

            li += 1
            if amount_val is None:
                continue
            if abs(amount_val) > MAX_PLAUSIBLE_AMOUNT:
                continue

            if name_text:
                rows.append((page_idx, current_customer, amount_val, line_text, source))
            elif other_words:
                if current_customer:
                    rows.append((page_idx, current_customer, amount_val, line_text, source))
            # reine Zahlen-Zeile ohne sonstigen Inhalt -> Zwischensumme, skip

    return rows


def extract_fondsfinanz(pages_words):
    """Bloc-Format von Fondsfinanz: 'VN Nachname, Vorname ...' gefolgt von
    einer Provisionsuebersicht und einer Zeile 'Vertragssumme EUR X'.

    Fondsfinanz behaelt zusaetzlich einen pauschalen, nicht kundenbezogenen
    'Stornoreserve'-Betrag als Sicherheit ein (siehe Deckblatt). Nach
    Rueckmeldung des Nutzers zaehlt fuer die Courtage-Erfassung die volle
    (versteuerte) Courtage inkl. dieses Einbehalts, nicht der um die Reserve
    gekuerzte Auszahlungsbetrag - der Einbehalt wird daher als eigene,
    ausdruecklich nicht-kundenspezifische Zeile wieder hinzugerechnet."""
    rows = []
    current_name = None
    storno_total = 0.0
    storno_line = None
    storno_page = None
    for page_idx, words, source in pages_words:
        if not words:
            continue
        lines = cluster_lines(words)
        for line in lines:
            toks = [w["text"] for w in line["words"]]
            if not toks:
                continue
            line_text = " ".join(toks)
            if toks[0] == "VN":
                if len(toks) >= 3 and toks[1].endswith(","):
                    current_name = toks[1] + " " + toks[2]
                elif len(toks) >= 2:
                    current_name = toks[1]
            elif toks[0] == "VP":
                continue
            elif "Vertragssumme" in toks:
                idx = toks.index("Vertragssumme")
                amt = None
                for t in toks[idx + 1:]:
                    amt = parse_amount(t)
                    if amt is not None:
                        break
                if amt is not None and current_name:
                    rows.append((page_idx, current_name, amt, line_text, source))
            elif toks[0] == "Stornoreserve":
                amt = parse_amount(toks[1]) if len(toks) > 1 else None
                if amt is not None and amt != 0:
                    storno_total += abs(amt)
                    storno_line = line_text
                    storno_page = page_idx

    if storno_total:
        rows.append((
            storno_page,
            "Stornoreserve-Einbehalt (nicht kundenspezifisch, siehe Deckblatt)",
            round(storno_total, 2),
            storno_line,
            "text",
        ))
    return rows


def extract_vn_summevertrag_block(pages_words):
    """Format wie Swiss Life 'Anhang Einzelaufstellung': Name und Betrag
    stehen auf derselben Zeile, z.B.
    'Vers.Nr. 9317437 VN/VP Knaus, Maximilian Referenznummer 74
    Summe Vertrag 82,03'."""
    rows = []
    for page_idx, words, source in pages_words:
        if not words:
            continue
        lines = cluster_lines(words)
        for line in lines:
            toks = [w["text"] for w in line["words"]]
            if "VN/VP" in toks:
                vn_idx = toks.index("VN/VP")
            elif "VN" in toks:
                vn_idx = toks.index("VN")
            else:
                continue
            if vn_idx + 1 >= len(toks):
                continue
            if toks[vn_idx + 1].endswith(",") and vn_idx + 2 < len(toks):
                name = toks[vn_idx + 1] + " " + toks[vn_idx + 2]
            else:
                name = toks[vn_idx + 1]

            amt = None
            for i in range(len(toks) - 2):
                if toks[i] == "Summe" and toks[i + 1] == "Vertrag":
                    amt = parse_amount(toks[i + 2])
                    break
            if amt is not None:
                line_text = " ".join(toks)
                rows.append((page_idx, name, amt, line_text, source))
    return rows


SWISS_LIFE_VSV_RE = re.compile(
    r"BeiträgezurVertrauensschadenversicherung\s+[\d.,]+\s+(-?[\d.,]+)\s+(-?[\d.,]+)"
)


def extract_swiss_life_vsv_deduction(pages_full_text):
    """Swiss Life behaelt pro Abrechnung einen pauschalen (nicht kunden-
    bezogenen) Beitrag zur Vertrauensschadenversicherung ein - anders als
    urspruenglich angenommen wird dieser NICHT separat von RH persoenlich
    getragen, sondern ist direkt von der Abschlussprovision abzuziehen
    (Nutzer-Rueckmeldung). Er steht explizit auf der "Abrechnungsuebersicht
    zum ..."-Seite (nicht auf der "Kumulierte Jahreswerte"-Seite, die
    dieselbe Zeile fuer das gesamte Jahr zeigt) als eigene Zeile
    "BeitraegezurVertrauensschadenversicherung <Gutschrift> <Belastung>
    <Saldo>", z.B. "... 0,00 -3,58 -3,58" - der Saldo-Wert wird als eigene,
    ausdruecklich nicht-kundenspezifische Abzugszeile zurueckgegeben."""
    for pidx, text in pages_full_text:
        if "SummenausaktuellerAbrechnung" not in text:
            continue
        m = SWISS_LIFE_VSV_RE.search(text)
        if m:
            saldo = parse_amount(m.group(2))
            if saldo:
                return pidx, saldo, m.group(0)
    return None


VHV_ADDR_RE = re.compile(r"^(.+?),\s*.+?,\s*\d{5}\s+\S")


def extract_vhv(pages_words):
    """VHV: 1-3 Buchungszeilen (Betrag als letztes Feld) stehen VOR der
    zugehoerigen Zeile 'Firma/Name, Strasse , PLZ Ort' - nicht danach wie
    bei den meisten anderen Versicherern. Die gepufferten Betraege werden
    rueckwirkend der folgenden Name+Adresse-Zeile zugeordnet."""
    rows = []
    for page_idx, words, source in pages_words:
        if not words:
            continue
        lines = cluster_lines(words)
        pending = []
        for line in lines:
            toks = [w["text"] for w in line["words"]]
            if not toks:
                continue
            line_text = " ".join(toks)
            m = VHV_ADDR_RE.match(line_text)
            if m:
                name = line_text.split(",")[0].strip()
                for amt, ltxt in pending:
                    rows.append((page_idx, name, amt, ltxt, source))
                pending = []
                continue
            if is_blacklisted_line(line_text.lower()):
                pending = []
                continue
            if len(toks) >= 6:
                amt = parse_amount(toks[-1])
                if amt is not None and abs(amt) <= MAX_PLAUSIBLE_AMOUNT:
                    pending.append((amt, line_text))
    return rows


def extract_sv_sparkasse(pages_words):
    """SV Sparkassenversicherung (OCR, kein Tabellenkopf): jede Buchungszeile
    beginnt mit dem Kundennamen, gefolgt von Codes/Daten, und endet mit einem
    Betrag gefolgt von einem 1-2-stelligen Flag-Zeichen (Ziffer oder durch
    OCR verfaelschter Buchstabe), z.B.
    'Hu, Hsiao-Lei 10 00 2 50090787898 BP ... 29,06 25,000 7,27 2'.
    Seite 0 (Anschreiben) und 1 (Aenderungsgrund-Legende) enthalten keine
    Buchungen und werden uebersprungen."""
    rows = []
    for page_idx, words, source in pages_words:
        if not words or page_idx < 2:
            continue
        lines = cluster_lines(words)
        for line in lines:
            toks = [w["text"] for w in line["words"] if "#" not in w["text"]]
            if len(toks) < 8:
                continue
            line_text = " ".join(toks)
            if is_blacklisted_line(line_text.lower()):
                continue
            name_toks = []
            i = 0
            while i < len(toks) and not re.fullmatch(r"[\d.,]+", toks[i]):
                name_toks.append(toks[i])
                i += 1
            if not name_toks or i >= len(toks) - 3:
                continue
            name = " ".join(name_toks).strip(" ,")
            if len(re.sub(r"[^A-Za-zÀ-ÿ]", "", name)) < 3:
                continue
            last = toks[-1]
            # der Betrag hat ein Komma (Dezimaltrenner), das nachgestellte
            # Flag-Zeichen nicht.
            amt_tok = last if "," in last else (toks[-2] if len(toks) >= 2 else last)
            amt = parse_amount(amt_tok)
            if amt is None or abs(amt) > MAX_PLAUSIBLE_AMOUNT:
                continue
            rows.append((page_idx, name, amt, line_text, source))
    return rows


ARAG_NAME_RE = re.compile(r"07(?!\d)\s*[:.\|]?\s*(.+?)\s+70\b")
ARAG_NUM_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")


def extract_arag(pdf):
    """ARAG-Einzelaufstellung (z.B. 'PROVISIONS-PRODUKTIONSABRECHNUNG'):
    gescanntes PDF, dessen Positionstabelle mit dem Standard-OCR-Pfad
    (get_words_for_page, 300dpi) zu schlecht erkannt wird. Hier wird pro
    Seite mit hoeherer Aufloesung (400dpi) und tabellenorientiertem OCR-Modus
    (--psm 6) als Fliesstext gelesen und zeilenweise per Regex geparst, statt
    ueber Wort-Positionen: 'AT-Nr 07 Nachname, Vorname 70 Beitrag Satz
    Provision [Provision wiederholt] Datum'.

    Storno-Zeilen (Beitrag mit Minus) tauchen manchmal als zweite Zeile mit
    demselben Namen und demselben (positiven) Provisionsbetrag auf, ohne
    dass die OCR das Minuszeichen zuverlaessig erkennt (schwankt je nach
    Aufloesung). Statt uns auf dieses Vorzeichen zu verlassen, wird daher
    jede Zeile, deren (normalisierter) Name UND Betrag mit der direkt
    vorherigen Zeile uebereinstimmen, als Storno der vorherigen Zeile
    behandelt und negativ gezaehlt - das entspricht der Beobachtung, dass
    beide Zeilen sich in der echten Abrechnung zu 0 aufheben."""
    rows = []
    prev_norm_key = None
    for page_idx, page in enumerate(pdf.pages):
        if len(page.chars) > 0:
            continue
        im = page.to_image(resolution=400).original
        text = pytesseract.image_to_string(im, lang="deu", config="--psm 6")

        for line in text.splitlines():
            if not re.search(r"07(?!\d)", line):
                continue
            m = ARAG_NAME_RE.search(line)
            if not m:
                continue
            name = re.sub(r"^[^A-Za-zÀ-ÿ]+", "", m.group(1)).strip(" .|_-;:")
            name_norm = re.sub(r"[^A-Za-zÀ-ÿ]", "", name).lower()
            if len(name_norm) < 3:
                continue
            nums = ARAG_NUM_RE.findall(line)
            if not nums:
                continue
            amt = parse_amount(nums[-1])
            if amt is None or abs(amt) > MAX_PLAUSIBLE_AMOUNT:
                continue

            norm_key = (name_norm, round(abs(amt), 2))
            if norm_key == prev_norm_key:
                # Storno-Gegenbuchung derselben Person/desselben Betrags.
                amt = -abs(amt)
                prev_norm_key = None
            else:
                prev_norm_key = norm_key
            rows.append((page_idx, name, amt, line.strip(), "ocr"))
    return rows


ALLIANZ_NUM_RE = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$")
ALLIANZ_CONTRACT_RE = re.compile(r"^AS-\d+$")
ALLIANZ_DIGITS_RE = re.compile(r"^\d+$")


def extract_allianz(pdf):
    """Allianz-Buchungsnote: Text pro Seite ist gespiegelt (siehe
    REVERSED_MARKERS/is_mirrored_text), UND die Tabelle hat zwei getrennte
    Betragsspalten "Belastung" (Storno/Ruckbuchung, negativ) und
    "Gutschrift" (Neuprovision, positiv). Beide Spalten koennen fuer
    dieselbe Vertragsnummer im selben Dokument vorkommen (z.B. Storno einer
    Rate UND eine neue Gutschrift fuer denselben Vertrag) - das ist KEIN
    Duplikat, siehe Nutzer-Rueckmeldung. Fruehere Versuche, die beide
    Vorkommen einfach zu addieren, ueberzaehlten deshalb systematisch.

    Wegen der Text-Spiegelung liefert pdfplumber verdrehte Koordinaten: die
    'top'-Achse verhaelt sich wie die horizontale (Spalten-)Achse, 'x0' wie
    die vertikale (Zeilen-)Achse. Die Spaltenzugehoerigkeit eines Betrags
    wird daher ueber seine 'top'-Position relativ zu den Kopfzeilen-Woertern
    'Belastung'/'Gutschrift' bestimmt. Jede Seite traegt zusaetzlich eine
    laufende Summenzeile ("Summe SACH-PROVISIONEN") ohne Vertragsnummer -
    wird ueber das Fehlen einer "AS-..."-Vertragsnummer ausgeschlossen.

    Verifiziert gegen den in jedem Dokument selbst aufgedruckten
    "Sach-Provisionen"-Endbetrag (Deckblatt) - exakte Uebereinstimmung bei
    allen 5 getesteten Juni-2026-Dateien."""
    rows = []
    for page_idx, page in enumerate(pdf.pages):
        words = page.extract_words()
        if not words:
            continue
        rwords = [{"text": w["text"][::-1], "top": w["top"], "x0": w["x0"]} for w in words]
        gut_top = next((w["top"] for w in rwords if w["text"] == "Gutschrift"), None)
        bel_top = next((w["top"] for w in rwords if w["text"] == "Belastung"), None)
        if gut_top is None or bel_top is None:
            continue  # Deckblatt oder sonstige Seite ohne SACH-PROVISIONEN-Tabelle
        boundary = (gut_top + bel_top) / 2

        groups = {}
        for w in rwords:
            key = round(w["x0"], 1)
            groups.setdefault(key, []).append(w)
        lines = [sorted(groups[k], key=lambda w: -w["top"]) for k in sorted(groups.keys())]

        for gi, line in enumerate(lines):
            texts = [w["text"] for w in line]
            contract = next((t for t in texts if ALLIANZ_CONTRACT_RE.match(t)), None)
            if not contract:
                continue
            amount_word = line[-1]
            if not ALLIANZ_NUM_RE.match(amount_word["text"]):
                continue
            sign = 1 if amount_word["top"] < boundary else -1
            amt = sign * parse_amount(amount_word["text"])

            # Kundenname steht in der jeweils naechsten Zeilengruppe, vor
            # dem ersten rein numerischen Token (Datev-Kennziffer, z.B.
            # "8072") - siehe Modul-Docstring.
            name = contract
            if gi + 1 < len(lines):
                name_tokens = []
                for t in (w["text"] for w in lines[gi + 1]):
                    if ALLIANZ_DIGITS_RE.match(t):
                        break
                    name_tokens.append(t)
                if name_tokens:
                    name = " ".join(name_tokens).strip()

            line_text = " ".join(texts)
            rows.append((page_idx, name, round(amt, 2), line_text, "text"))
    return rows


AXA_VN_MARKER = "VN:"


def extract_axa(pdf):
    """AXA-Einzelnachweis zum Kontoauszug: Text pro Seite ist gespiegelt
    (siehe REVERSED_MARKERS/is_mirrored_text), Layout wie bei Allianz mit
    vertauschten Koordinatenachsen (siehe extract_allianz-Docstring).

    Jeder Kundenblock beginnt mit einer Zeile 'VN: Nachname, Vorname ...'
    und kann mehrere Sparten-/Vertragszeilen enthalten; die letzte Zeile
    des Blocks ist eine mit Sternchen maskierte Summenzeile, die auf 'VD'
    endet und den Netto-Provisionsbetrag des ganzen Blocks traegt (Format:
    '<Betrag> <H|S> <Betrag> <H|S> VD' - Provisions- und Abrechnungsbetrag
    sind identisch). Nur diese Summenzeile wird verwendet (nicht die
    einzelnen Sparten-Teilbetraege), sonst wuerde doppelt gezaehlt.

    Auf der letzten Seite folgt zusaetzlich ein 'ALLGEMEINE UMSAETZE
    VD-KONTO'-Abschnitt mit kontobezogenen (nicht kundenbezogenen)
    Buchungen wie 'Kontoausgleich' oder 'Umbuchung' - diese enden zwar
    ebenfalls auf 'VD', haben aber keinen vorausgehenden echten
    'VN:'-Kundenblock und werden daher ueber current_name=None
    ausgeschlossen. Verifiziert: die Summe der Kundenbloecke stimmt bei
    allen 3 getesteten Juni-2026-Dateien exakt mit der im PDF selbst
    aufgedruckten 'Betreu.-prov.'-Kontrollsumme ueberein."""
    rows = []
    for page_idx, page in enumerate(pdf.pages):
        words = page.extract_words()
        if not words:
            continue
        rwords = [{"text": w["text"][::-1], "top": w["top"], "x0": w["x0"]} for w in words]
        groups = {}
        for w in rwords:
            key = round(w["x0"], 1)
            groups.setdefault(key, []).append(w)
        lines = [sorted(groups[k], key=lambda w: -w["top"]) for k in sorted(groups.keys())]

        current_name = None
        for line in lines:
            texts = [w["text"] for w in line]
            # Nur der EXAKTE Token "VN:" zaehlt als Start eines echten
            # Kundenblocks. In der "ALLGEMEINE UMSAETZE"-Sektion (letzte
            # Seite) kommt gelegentlich ein mit dem Namen verschmolzenes
            # "VN:Name" als blosser Referenzhinweis auf einer allgemeinen,
            # nicht kundenspezifischen Kontobuchung vor (z.B. "Kontoaus-
            # gleich") - wuerde bei Erkennung faelschlich current_name
            # setzen und diese Buchung so einem Kunden zurechnen.
            if AXA_VN_MARKER in texts:
                vn_idx = texts.index(AXA_VN_MARKER)
                name_toks = texts[vn_idx + 1:vn_idx + 3]
                current_name = " ".join(name_toks).replace(",", "").strip()
                continue
            if texts and texts[-1] == "VD" and len(texts) >= 4:
                amt_tok, flag_tok = texts[-3], texts[-2]
                if ALLIANZ_NUM_RE.match(amt_tok) and flag_tok in ("H", "S") and current_name:
                    sign = 1 if flag_tok == "H" else -1
                    amt = sign * parse_amount(amt_tok)
                    line_text = " ".join(texts)
                    rows.append((page_idx, current_name, round(amt, 2), line_text, "text"))
    return rows


GOTHAER_CONTRACT_RE = re.compile(r"^\d[\d.\-/]{5,}$")
GOTHAER_NAME_RE = re.compile(r"^([A-ZÀ-Ü][a-zà-ÿß]+),([A-ZÀ-Ü][a-zà-ÿß]+)")


def extract_gothaer(pdf):
    """Gothaer (Allgemeine + Leben): gespiegelter Text wie bei Allianz/AXA
    (siehe extract_allianz-Docstring fuer die vertauschten Koordinaten).

    Jede Buchungszeile beginnt mit einer Vertrags-/Vers.-Nr. (z.B.
    '38.038.906337' oder '64-818940-60/0000') und endet mit dem C/P-Betrag
    (Courtage/Provision); der zugehoerige Kundenname ('Nachname,Vorname',
    ohne Leerzeichen zusammengeschrieben) steht in der jeweils naechsten
    Zeilengruppe. Verifiziert gegen die auf dem Deckblatt aufgedruckte
    "Endsumme" (Gunsten-Spalte) - exakte Uebereinstimmung bei beiden
    getesteten Juni-2026-Dateien (14,21 EUR / 1,50 EUR)."""
    rows = []
    for page_idx, page in enumerate(pdf.pages):
        words = page.extract_words()
        if not words:
            continue
        rwords = [{"text": w["text"][::-1], "top": w["top"], "x0": w["x0"]} for w in words]
        groups = {}
        for w in rwords:
            key = round(w["x0"], 1)
            groups.setdefault(key, []).append(w)
        keys = sorted(groups.keys())
        lines = [sorted(groups[k], key=lambda w: -w["top"]) for k in keys]

        for gi, line in enumerate(lines):
            texts = [w["text"] for w in line]
            if not texts:
                continue
            if GOTHAER_CONTRACT_RE.match(texts[0]) and ALLIANZ_NUM_RE.match(texts[-1]):
                amt = parse_amount(texts[-1])
                name = None
                if gi + 1 < len(lines):
                    ntexts = [w["text"] for w in lines[gi + 1]]
                    if ntexts:
                        m = GOTHAER_NAME_RE.match(ntexts[0])
                        if m:
                            name = f"{m.group(1)} {m.group(2)}"
                if name:
                    line_text = " ".join(texts)
                    rows.append((page_idx, name, round(amt, 2), line_text, "text"))
    return rows


CONTINENTALE_NAME_RE = re.compile(r"([A-ZÀ-Ü][a-zà-ÿß]+-[A-ZÀ-Ü][a-zà-ÿß]+)")
CONTINENTALE_CLEAN_AMT_RE = re.compile(r"\+(\d{1,3}(?:\.\d{3})*,\d{2})")
CONTINENTALE_FALLBACK_AMT_RE = re.compile(r"[+H]\s*(\d{2,4})\b")
CONTINENTALE_SALDO_RE = re.compile(r"Neuer Saldo:?\s*(-?[\d.,]+)")


def _continentale_candidates(lines):
    """Liefert je Namenszeile (name, sauberer_Betrag_oder_None,
    Rueckfall_Betrag_oder_None, Zeile) - siehe extract_continentale()."""
    out = []
    for line in lines:
        if "P__" not in line and "P___" not in line:
            continue
        m = CONTINENTALE_NAME_RE.search(line)
        if not m:
            continue
        rest = line[m.end():]
        clean = None
        cm = CONTINENTALE_CLEAN_AMT_RE.search(rest)
        if cm:
            clean = parse_amount(cm.group(1))
        fallback = None
        fm = CONTINENTALE_FALLBACK_AMT_RE.search(rest)
        if fm:
            digits = fm.group(1)
            # Bei manchen Zeilen faellt der OCR das Komma weg (z.B. "021"
            # statt "0,21") - die letzten beiden Ziffern sind dann die
            # Nachkommastellen (empirisch verifiziert: ergibt zusammen mit
            # den sauber gelesenen Zeilen exakt den aufgedruckten "Neuer
            # Saldo"-Betrag).
            fallback = parse_amount(f"{digits[:-2] or '0'},{digits[-2:]}")
        out.append((m.group(1), clean, fallback, line.strip()))
    return out


def extract_continentale(pdf):
    """Continentale 'Provisionsnote Einzelergebnisse' (gescannt): sehr
    dichte, mehrspaltige Tabelle mit viel Rahmen-/Trennzeichen-Bildrauschen,
    bei der einzelne Betragswerte je nach OCR-Aufloesung mal sauber
    ("+19,45"), mal mit verlorenem Komma ("+502" statt "+5,02") oder mit
    Buchstaben-Ersatzzeichen ("H021" statt "+0,21", vermutlich ein
    OCR-Fehllesen eines "+"-Symbols) gelesen werden - und das nicht
    konsistent bei derselben Aufloesung fuer alle Zeilen.

    Deshalb: OCR bei mehreren Aufloesungen (400/500/600dpi) durchfuehren,
    je Namenszeile ueber die Reihenfolge (nicht den exakten Text, der pro
    Aufloesung variiert) zuordnen und den saubersten verfuegbaren Treffer
    nehmen (zuerst ein Betrag mit Komma, sonst die Ziffern-Rueckfalllogik).

    Da diese Rueckfalllogik unvermeidbar heuristisch ist, wird das Ergebnis
    IMMER gegen den auf der Kontoauszug-Seite aufgedruckten "Neuer Saldo"-
    Betrag geprueft (siehe process_file()): stimmt die Summe nicht exakt
    ueberein, werden die Zeilen verworfen und die Datei faellt automatisch
    auf manuelle Pruefung zurueck, statt falsche Kundenzuordnungen zu
    riskieren - passend zum Grundsatz "Summe der Einzelpositionen muss immer
    dem Gesamtsaldo entsprechen, im Zweifel Rueckmeldung statt Raten"."""
    rows = []
    target_total = None
    for pidx, page in enumerate(pdf.pages):
        if len(page.chars) > 0:
            continue
        quick_text = pytesseract.image_to_string(page.to_image(resolution=200).original, lang="deu")
        if target_total is None and "Neuer Saldo" in quick_text:
            m = CONTINENTALE_SALDO_RE.search(quick_text)
            if m:
                target_total = parse_amount(m.group(1))
        if "Einzelergebnisse" not in quick_text:
            continue

        per_resolution = []
        for res in (400, 500, 600):
            im = page.to_image(resolution=res).original
            text = pytesseract.image_to_string(im, lang="deu", config="--psm 6")
            per_resolution.append(_continentale_candidates(text.splitlines()))

        n = max((len(c) for c in per_resolution), default=0)
        for i in range(n):
            entries = [c[i] for c in per_resolution if i < len(c)]
            if not entries:
                continue
            name = entries[0][0]
            amt, chosen_line = None, entries[0][3]
            for _, clean, _fallback, line in entries:
                if clean is not None:
                    amt, chosen_line = clean, line
                    break
            if amt is None:
                for _, _clean, fallback, line in entries:
                    if fallback is not None:
                        amt, chosen_line = fallback, line
                        break
            if amt is not None:
                rows.append((pidx, name, round(amt, 2), chosen_line, "ocr"))
    return rows, target_total


def extract_alte_leipziger(pages_words):
    """Alte Leipziger (gescannt, OCR): sauber gescannte, saubere Tabelle
    ('Abrechnungskonto (Faellige Verguetungen)'), aber mit zwei Eigenheiten
    gegenueber der generischen Tabellen-Engine:

    1) Der Betrag ('Gutschrift/Belastung') steht auf der ERSTEN Zeile eines
       mehrzeiligen Kundenblocks (zusammen mit der Vertrags-/Versicherungs-
       nummer), der lesbare Kundenname aber oft erst auf der 2./3.
       Folgezeile (Firmenname mit Zeilenumbruch) - umgekehrte Reihenfolge
       zur generischen Engine. Betraege werden daher pro Block gesammelt
       und erst bei 'Zwischensumme' (Blockende) dem inzwischen bekannten
       Namen zugeordnet.
    2) Die Kopfzeile 'Name VN' ist zweiteilig; die generische
       Spaltenbreiten-Erkennung haelt das zweite Wort 'VN' faelschlich fuer
       den Beginn der naechsten Spalte und schneidet lange Firmennamen nach
       dem ersten Wort ab - die Namensspalte wird daher manuell bis zum
       bekannten Beginn der naechsten Spalte ('Ausloeser', x0~128) erweitert.

    Nur Seiten mit "Abrechnungskonto" im Text werden verarbeitet (die
    anderen Seiten sind Deckblatt/Inhaltsverzeichnis/Rechtshinweise ohne
    Buchungsdaten). Verifiziert: Summe je Teilabrechnung entspricht exakt
    dem Betrag der zugehoerigen "Zahlungsausgang"-Zeile (Kontrollsumme im
    PDF selbst) - 7.870,01 EUR fuer die 3 Juni-2026-Dateien zusammen."""
    rows = []
    columns = None
    block_name_parts = []
    block_amounts = []

    def flush(page_idx, source):
        nonlocal block_name_parts, block_amounts
        name_full = ""
        for frag in block_name_parts:
            if name_full.endswith("-"):
                name_full += frag
            else:
                name_full = (name_full + " " + frag).strip()
        if name_full:
            for amt, ltxt in block_amounts:
                rows.append((page_idx, name_full, amt, ltxt, source))
        block_name_parts = []
        block_amounts = []

    for page_idx, words, source in pages_words:
        flush(page_idx, source)
        columns = None
        page_text = " ".join(w["text"] for w in words).lower()
        if not words or "abrechnungskonto" not in page_text:
            continue
        lines = cluster_lines(words)
        li, n = 0, len(lines)
        while li < n:
            # Kopfzeile ist immer genau 4 Zeilen lang (die 4. Zeile
            # "Herkunftsvermittler (VVNR)" traegt sonst faelschlich zur
            # erkannten Spaltenbreite bei).
            header_found = _header_from_span(lines, li, 4)
            if header_found:
                flush(page_idx, source)
                columns, next_li = header_found
                columns = dict(columns)
                columns["name_range"] = (columns["name_range"][0], 120)
                li = next_li
                continue
            if columns is None:
                li += 1
                continue

            line = lines[li]
            line_text = " ".join(w["text"] for w in line["words"])
            line_lower = line_text.lower()
            if "vvnr" in line_lower or "herkunftsvermittler" in line_lower:
                # Rest-Echo der 4. Kopfzeilen-Zeile, das trotz Kopfzeilen-
                # Erkennung gelegentlich noch als Datenzeile ankommt.
                li += 1
                continue
            if is_blacklisted_line(line_lower):
                if "zwischensumme" in line_lower:
                    flush(page_idx, source)
                elif "kontenstand" in line_lower:
                    # Tabellenende: alles danach ist Fusszeilen-Boilerplate
                    # (IBAN/Postbank/Direktion/...), keine Buchungsdaten mehr.
                    flush(page_idx, source)
                    break
                li += 1
                continue

            name_l, name_r = columns["name_range"]
            amt_l, amt_r = columns["amount_range"]
            name_words = [w for w in line["words"] if name_l <= w["x0"] < name_r]
            amount_words = [w for w in line["words"] if amt_l <= w["x0"] < amt_r]
            name_text = " ".join(w["text"] for w in name_words).strip(" ,")
            letter_runs = re.findall(r"[A-Za-zÀ-ÿ]+", name_text)
            longest_run = max((len(r) for r in letter_runs), default=0)
            amount_val = None
            for w in amount_words:
                v = parse_amount(w["text"])
                if v is not None:
                    amount_val = v
                    break

            li += 1
            if longest_run >= 3:
                block_name_parts.append(name_text)
            if amount_val is not None and abs(amount_val) <= MAX_PLAUSIBLE_AMOUNT:
                block_amounts.append((amount_val, line_text))
    return rows


BARMENIA_NAME_RE = re.compile(r"^([A-ZÀ-Ü][A-Za-zà-ÿ\-]+),\s*([A-ZÀ-Ü][A-Za-zà-ÿ\-]+)")
BARMENIA_NUM_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")


def extract_barmenia(pdf):
    """Barmenia/Gothaer 'Vergluetungsnachweis Abschlussverguetungen'
    (gescannt): der Standard-OCR-Pfad (300dpi) verliert auf den
    Detailseiten regelmaessig die eigentlichen Betragsspalten (zu dicht/
    klein gedruckte Tabelle) - hoehere Aufloesung (400dpi) mit
    tabellenorientiertem OCR-Modus (--psm 6) als Fliesstext liest sie
    zuverlaessig. Pro Kundenzeile (Muster 'Nachname, Vorname ...') steht
    der Verguetungsbetrag doppelt hintereinander (Verguetungsbetrag =
    Abgerechneter Betrag, sofern kein Storno/keine Proration) - dieses
    Zahlenpaar wird als der gesuchte Betrag genommen. Nur Seiten mit
    'Verguetungsnachweis' im Text enthalten Kundenpositionen (Deckblatt/
    Kontoauszug/Gesamtuebersicht-Seiten werden uebersprungen)."""
    rows = []
    for pidx, page in enumerate(pdf.pages):
        if len(page.chars) > 0:
            continue
        quick_text = pytesseract.image_to_string(page.to_image(resolution=200).original, lang="deu")
        if "vergütungsnachweis" not in quick_text.lower():
            continue
        im = page.to_image(resolution=400).original
        text = pytesseract.image_to_string(im, lang="deu", config="--psm 6")
        for line in text.splitlines():
            m = BARMENIA_NAME_RE.match(line.strip())
            if not m:
                continue
            name = f"{m.group(1)} {m.group(2)}"
            amounts = [parse_amount(a) for a in BARMENIA_NUM_RE.findall(line)]
            amt = None
            for i in range(len(amounts) - 1):
                if amounts[i] == amounts[i + 1] and amounts[i]:
                    amt = amounts[i]
                    break
            if amt is not None:
                rows.append((pidx, name, amt, line.strip(), "ocr"))
    return rows


VEMA_CSV_AMOUNT_COL = "Betrag"


def extract_vema_csv(path):
    """VEMA-Pool liefert die Kundenpositionen NICHT im PDF (das ist ein
    reiner Sammelbeleg, siehe README), sondern separat als CSV-Export
    ('VEMA-Poolabrechnung-N.csv'). Spalten (Semikolon-getrennt):
    VEMAintern;VN Anrede;VN Vorname;VN Nachname;Vertragsnummer;
    Gesellschaft;Sparte;Faelligkeit;Provisionsbasis;Courtage;
    Provisionsart;Einbehalt;Betrag;...

    'Courtage' ist der Brutto-Courtagebetrag, 'Betrag' der Betrag NACH
    Abzug des VEMA-Poolbeitrags (siehe Spalte 'Einbehalt', typischerweise
    10%). Anders als z.B. bei Fondsfinanz's Stornoreserve (temporaere
    Sicherheit, siehe extract_fondsfinanz) ist der VEMA-Einbehalt ein
    dauerhafter Pool-Verwaltungsbeitrag, der SSH nie zufliesst - massgeblich
    ist daher 'Betrag' (netto), nicht 'Courtage' (brutto)."""
    rows = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        if VEMA_CSV_AMOUNT_COL not in (reader.fieldnames or []):
            return rows
        for row in reader:
            anrede = (row.get("VN Anrede") or "").strip()
            vorname = (row.get("VN Vorname") or "").strip()
            nachname = (row.get("VN Nachname") or "").strip()
            if anrede == "Firma" or not vorname:
                name = nachname
            else:
                name = f"{nachname}, {vorname}"
            betrag_s = (row.get(VEMA_CSV_AMOUNT_COL) or "").strip()
            if not betrag_s or not name:
                continue
            betrag = parse_amount(betrag_s)
            if betrag is None:
                continue
            raw_line = ";".join(f"{k}={v}" for k, v in row.items() if v)
            rows.append((0, name, round(betrag, 2), raw_line, "csv"))
    return rows


def find_total_in_text(full_text):
    """Sucht best-effort einen Gesamt-/Endbetrag im Freitext (fuer
    Sammelbelege ohne Kundendetail, zur Kontrolle)."""
    candidates = []
    for pat in TOTAL_LINE_PATTERNS:
        for m in re.finditer(pat, full_text, re.IGNORECASE | re.MULTILINE):
            val = parse_amount(m.group(1))
            if val is not None:
                candidates.append(val)
    if candidates:
        return candidates[-1]
    return None


def is_mirrored_text(full_text):
    return any(marker in full_text for marker in REVERSED_MARKERS)


# ---------------------------------------------------------------------------
# Kontoauszug-Abgleich (VR Bank Rhein-Neckar): erkennt Zahlungseingaenge ohne
# zugehoerige Abrechnungs-PDF fuer den Monat.
# ---------------------------------------------------------------------------

BANK_TX_START_RE = re.compile(
    r"^(\d{2}\.\d{2})\.\s+(\d{2}\.\d{2})\.\s+(.+?)\s+PN:\d+\s+(-?[\d.,]+)\s+([HS])\s*$"
)
# Wiederkehrende Brief-/Fusszeilen (Bankname, Adresse, Kontostand-Zusammen-
# fassung, Seitenumbruch-Hinweise) - koennen an beliebiger Stelle in der
# Zeile stehen (z.B. "K00062862 Bitte beachten Sie ..."), daher ohne
# Zeilenanfang-Anker.
BANK_BOILERPLATE_RE = re.compile(
    r"(VR Bank Rhein-Neckar|www\.vrbank\.de|Kontokorrent|EUR-Konto Kontonummer|"
    r"Postfach \d|neuer Kontostand|Gesamtumsatz:|Ihr Kreditrahmen|\*{5,}|"
    r"Augustaanlage|Ihr Berater:|Telefon: 0|E-Mail:|Bu-Tag Wert Vorgang|"
    r"alter Kontostand|Übertrag (auf|von) Blatt|Bitte beachten Sie|"
    r"Kontoauszug Nr\.|erstellt am)"
)
# Zeilen, die NUR aus einem dieser Fragmente bestehen (Absender-/Empfaenger-
# Adressblock von SSH selbst, wiederholt sich auf jeder Seite) - werden nur
# bei exakter Volltreffer-Gleichheit gefiltert, damit ein echter Absendername
# mit z.B. "GmbH & Co.KG" als Namensbestandteil nicht faelschlich verworfen
# wird.
BANK_BOILERPLATE_WHOLE_LINE_RE = re.compile(
    r"^(K\d{8}|\d{4}|000|0522|SSH Versicherungsmakler|GmbH & Co\.KG|"
    r"Anne-Frank-Str\. 8|\d{5} Viernheim)$"
)

# Bekannte Absender-/Verwendungszweck-Stichworte je Versicherer-"Familie"
# (normalisiert: klein geschrieben, Umlaute uebersetzt, alles ausser a-z0-9
# entfernt - siehe normalize_for_match()). Ein Versicherer-Dateiname wird
# einer Familie zugeordnet, wenn sein normalisierter Name mit dem
# normalisierten Familienschluessel beginnt (siehe insurer_bank_family()) -
# so fallen z.B. "R+V-VM-076965" und "R+V-VM-214574" beide unter "r+v".
# Diese Liste ist zwangslaeufig unvollstaendig (neue Versicherer, geaenderte
# Bank-Bezeichnungen) - unbekannte Zahlungseingaenge werden deshalb nicht
# stillschweigend ignoriert, sondern explizit als "nicht zugeordnet"
# ausgewiesen (siehe reconcile_bank_credits()).
INSURER_BANK_ALIASES = {
    "allianz": ["allianz"],
    "axa": ["axa"],
    "gothaer-allgemeine": ["gothaerallgemeine"],
    "gothaer-leben": ["gothaerleben", "gothaerlebensversicherung"],
    "fondsfinanz": ["fondsfinanz"],
    "r+v": ["rvallgemeine", "zentralesdirektinkasso"],
    "mannheimer": ["mannheimer"],
    "alte-leipziger": ["alteleipziger"],
    "vema-pool": ["vemaversicherungsmaklergenossenschaft"],
    "wuerttembergische": ["wuerttversicherg", "wuerttembergische"],
    "hiscox": ["hiscox"],
    "deurag": ["deurag"],
    "signal-iduna": ["signaliduna"],
    "hdi-leben": ["hdileben"],
    "hdi": ["hdiversicherung"],
    "swiss-life": ["swisslife"],
    "vhv": ["vhv"],
    "itzehoer": ["itzehoerversvereinag", "itzehoer"],
    "markel": ["markel"],
    "haftpflichtkasse": ["haftpflichtkasse"],
    "herzenssache": ["herzenssache"],
    "sv-sparkassenversicherung": ["svholding", "svsparkassen"],
    "aig": ["aigeurope"],
    "amex-pool": ["amex", "qualitypool"],
    "interrisk": ["interrisk"],
    "dialog": ["dialogversicherung"],
    "arag": ["aragse"],
    "auxilia": ["kraftfahrerschutz"],
    "concordia": ["concordia"],
    "bsg": ["barmenia"],
}


def normalize_for_match(s):
    """Wie normalize(), zusaetzlich werden alle Nicht-alphanumerischen
    Zeichen entfernt (Leerzeichen, Punkte, Bindestriche, '+', ...) - macht
    Versicherer-Dateinamen und Bank-Freitext direkt vergleichbar."""
    return re.sub(r"[^a-z0-9]", "", normalize(s))


def insurer_bank_family(insurer_key):
    """Ordnet einen Versicherer-Dateinamen (z.B. 'R+V-VM-076965') seiner
    Bank-Abgleichs-'Familie' zu (z.B. 'r+v'). Unbekannte Versicherer liefern
    ihren eigenen normalisierten Namen zurueck (kein Fehler - wird beim
    Abgleich dann einfach nicht in INSURER_BANK_ALIASES gefunden)."""
    key_norm = normalize_for_match(insurer_key)
    for family in sorted(INSURER_BANK_ALIASES, key=len, reverse=True):
        if key_norm.startswith(normalize_for_match(family)):
            return family
    return key_norm


# ---------------------------------------------------------------------------
# Kunde -> Betreuer-Zuordnung (CRM-Export "Betreuer.xlsx")
# ---------------------------------------------------------------------------

# Andreas Selle hat selbst keinen CRM-Account - seine Kunden werden dort
# unter "Calvin Keinarth" gefuehrt (Nutzer-Angabe). Umsatz wird trotzdem
# Andreas Selle zugerechnet, nicht Calvin.
BETREUER_ALIASES = {"calvin keinarth": "Andreas Selle"}
PARTNER_NAMES = ["Robin Heckmann", "Tim Selle", "Andreas Selle"]
# Farbcodes je Partner fuer die farbige Markierung in Excel/PDF (Nutzer-
# Vorgabe: Rot=RH, Gelb=TS, Blau=AS).
PARTNER_COLORS_XLSX = {
    "Robin Heckmann": "FFC7CE",
    "Tim Selle": "FFEB9C",
    "Andreas Selle": "BDD7EE",
}
PARTNER_COLORS_PDF = {
    "Robin Heckmann": (200, 30, 30),
    "Tim Selle": (170, 140, 0),
    "Andreas Selle": (30, 80, 170),
}
# Kraeftigere Variante nur fuer die kleinen Legenden-Farbkaestchen (dort
# spielt Lesbarkeit von Text keine Rolle, ein sattes Gelb ist als "Gelb"
# aber deutlich eindeutiger erkennbar als das gedeckte PARTNER_COLORS_PDF-Gelb,
# das auf weissem Hintergrund lesbar bleiben muss).
PARTNER_SWATCH_COLORS_PDF = {
    "Robin Heckmann": (230, 50, 50),
    "Tim Selle": (255, 200, 0),
    "Andreas Selle": (50, 110, 220),
}

# Manuelle Korrekturen fuer Kunden, die match_betreuer() ueber die
# Betreuer.xlsx nicht (mehr) zuverlaessig finden kann - z.B. weil der Name
# im Quell-PDF stark abgekuerzt/abweichend ist, weil es sich um einen
# Handelsnamen statt des CRM-Firmennamens handelt, oder weil sich der
# Nachname zwischenzeitlich geaendert hat (Heirat) und daher keine noch so
# lockere Schreibvarianz je zum CRM-Namen passen wuerde. Vom Nutzer per
# Hand bestaetigt (Stand Juli 2026) - Key ist der ORIGINALE, im PDF
# auftauchende Kundenname (wird ueber normalize_for_match() abgeglichen,
# siehe apply_betreuer()). Bei neuen Faellen einfach ergaenzen; langfristig
# sauberer ist es, die Ursache direkt in der Betreuer.xlsx zu pflegen
# (z.B. Namensaenderung/Status "inaktiv" bei Rosen Adrian).
MANUAL_BETREUER_OVERRIDES_RAW = {
    "HVP": "Robin Heckmann",  # High Voltage Projects GmbH
    "Brötzmann Tanja": "Robin Heckmann",
    "Heckmann, Oliver": "Robin Heckmann",
    "Aghaie Pour": "Robin Heckmann",  # Hamid Aghaiepour
    "Shalamanova Snezhanka Dan": "Tim Selle",  # Snezhana Danailova Shalamanova
    "Calhanoglu Muhammed": "Robin Heckmann",
    "Rieker Anagbogu": "Robin Heckmann",  # Elvira Rieker Anagbogu
    "Rosen Adrian": "Robin Heckmann",  # in CRM als inaktiv markiert, aber aktiv
    "Dos Santos Card": "Robin Heckmann",  # Paulo Dos Santos Cardoso
    "Die Peperonis Heddesh": "Robin Heckmann",  # Die Peperonis Heddesheim Inh. Tuerkmen & Dornsbach
    "Pension Wohlgelegen": "Tim Selle",  # Pension Wohlgelegen Natascha San Miguel Teran
    "Turbo team Hatice Sengönül": "Robin Heckmann",  # Turbo Team Umzuege & Transporte
    "Hoffmann Holding Gmb": "Tim Selle",  # Hoffmann Holding GmbH
    "Gutperle & Czech": "Andreas Selle",  # Gutperle & Czech Projektentwicklungs GmbH
    "Katzir-Shimo": "Robin Heckmann",  # Alon Katzir-Shimoni
    "Saßmann-Herz": "Robin Heckmann",  # Gunther Saßmann
    "Van der": "Tim Selle",  # Karin Van der Raaij
    "Stasiak, Alexander": "Tim Selle",  # jetzt Alexander Weickel (Heirat)
    "Erbrecht, Felix": "Robin Heckmann",  # jetzt Felix Aleman (Heirat)
}

# Kennzeichen fuer nicht-kundenspezifische Sammelabzuege innerhalb einer
# Abrechnung (z.B. Fondsfinanz-Stornoreserve, Swiss-Life-VSV-Beitrag, siehe
# extract_swiss_life_vsv_deduction()) - werden in apply_betreuer() separat
# behandelt (Zuordnung zur groessten Kundenposition derselben Datei), nicht
# ueber die normale Namenszuordnung.
COLLECTIVE_DEDUCTION_MARKER = "nicht kundenspezifisch"

GEB_RE = re.compile(r"\(geb\.?[^)]*\)|\bgeb\.?:?\s.*$", re.IGNORECASE)
LEGAL_FORM_SUFFIXES = ["gmbhcokg", "gmbh", "cokg", "ug", "ev", "ag", "se", "kg", "ohg", "gbr"]


def _strip_geb(s):
    """Entfernt Geburtsname-Zusaetze wie '(geb. Bock)' oder 'geb.: Pleli'
    aus einem CRM-Namen, damit z.B. 'Dorn (geb. Bock)' auf 'Dorn' passt."""
    if not s:
        return s
    return GEB_RE.sub("", str(s)).strip()


def _strip_legal_form(s):
    for suf in LEGAL_FORM_SUFFIXES:
        if s.endswith(suf) and len(s) > len(suf):
            return s[: -len(suf)]
    return s


def _levenshtein_le1(a, b):
    """True, wenn a und b sich in hoechstens einem Zeichen unterscheiden
    (Einfuegen/Loeschen/Ersetzen) - fuer sehr enge Tippfehler-Varianten wie
    'Philip'/'Phillip'. Bewusst kein allgemeiner Fuzzy-Match (Risiko, zwei
    verschiedene Personen zu verwechseln), siehe _names_match()."""
    if abs(len(a) - len(b)) > 1:
        return False
    if a == b:
        return True
    # gleiche Laenge: genau ein ersetztes Zeichen erlaubt
    if len(a) == len(b):
        diffs = sum(1 for x, y in zip(a, b) if x != y)
        return diffs <= 1
    # Laengendifferenz 1: ein eingefuegtes/geloeschtes Zeichen erlaubt
    shorter, longer = (a, b) if len(a) < len(b) else (b, a)
    for i in range(len(longer)):
        if longer[:i] + longer[i + 1:] == shorter:
            return True
    return False


def _names_match(a, b):
    """Vergleicht zwei bereits normalisierte Vornamen. Exakt gleich,
    ODER einer ist Praefix des anderen (typisch fuer im Quell-PDF
    abgeschnittene Namen, z.B. 'wolfgan' vs 'wolfgang', 'seb' vs
    'sebastian' - siehe Nutzer-Bestaetigung), ODER Levenshtein-Distanz <= 1
    bei Namen ab 4 Zeichen (enge Schreibvarianten wie 'philip'/'phillip').
    Der Nachname muss IMMER exakt uebereinstimmen (siehe match_betreuer) -
    diese Unschaerfe gilt nur fuer den Vornamen, um keine unterschiedlichen
    Personen zu verwechseln."""
    if not a or not b:
        return False
    if a == b:
        return True
    if len(a) >= 3 and len(b) >= 3 and (a.startswith(b) or b.startswith(a)):
        return True
    if len(a) >= 4 and len(b) >= 4 and _levenshtein_le1(a, b):
        return True
    return False


def load_betreuer_lookup(path):
    """Laedt den CRM-Export 'Betreuer.xlsx' (Blaetter 'Privatkunden' mit
    Anrede/Vorname/Nachname/Kundenverantwortliche und 'Firmenkunden' mit
    u.a. Firmenname/Kundenverantwortliche) und baut daraus Nachschlage-
    Strukturen fuer match_betreuer(). 'Calvin Keinarth' wird auf 'Andreas
    Selle' umgemappt (siehe BETREUER_ALIASES)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    by_surname = {}  # normalisierter Nachname -> [(normalisierter Vorname, Betreuer)]
    people = []  # flache Liste (normalisierter Nachname, Vorname, Betreuer) - fuer Fuzzy-Nachname-Scan
    companies = {}  # normalisierter (Rechtsform-bereinigter) Firmenname -> Betreuer

    if "Privatkunden" in wb.sheetnames:
        ws = wb["Privatkunden"]
        header = [c.value for c in ws[1]]
        if "Vorname" in header and "Nachname" in header and "Kundenverantwortliche" in header:
            i_vor, i_nach, i_bet = (header.index(k) for k in
                                     ("Vorname", "Nachname", "Kundenverantwortliche"))
            for row in ws.iter_rows(min_row=2, values_only=True):
                vor, nach, bet = row[i_vor], row[i_nach], row[i_bet]
                if not bet or not nach:
                    continue
                bet = BETREUER_ALIASES.get(normalize(str(bet)), bet)
                nnach = normalize_for_match(_strip_geb(nach))
                nvor = normalize_for_match(_strip_geb(vor))
                if nnach:
                    by_surname.setdefault(nnach, []).append((nvor, bet))
                    people.append((nnach, nvor, bet))

    if "Firmenkunden" in wb.sheetnames:
        ws = wb["Firmenkunden"]
        header = [c.value for c in ws[1]]
        if "Firmenname" in header and "Kundenverantwortliche" in header:
            i_firma, i_bet = header.index("Firmenname"), header.index("Kundenverantwortliche")
            for row in ws.iter_rows(min_row=2, values_only=True):
                firma, bet = row[i_firma], row[i_bet]
                if not bet or not firma:
                    continue
                bet = BETREUER_ALIASES.get(normalize(str(bet)), bet)
                key = _strip_legal_form(normalize_for_match(firma))
                if key:
                    companies[key] = bet

    return {"by_surname": by_surname, "people": people, "companies": companies}


def match_betreuer(kunde, lookup):
    """Ordnet einen extrahierten Kundennamen (z.B. 'Nachname, Vorname',
    'Vorname Nachname' oder ein Firmenname) einem Betreuer aus dem CRM-
    Export zu (siehe load_betreuer_lookup()).

    Geprueft wird stufenweise, von streng zu locker (exakt vor unscharf,
    Nachname vor Vorname), und bei jeder Stufe gilt: sobald genau EIN
    Betreuer zu den Kandidaten passt, ist das Ergebnis "eindeutig genug"
    und wird zurueckgegeben (Nutzer-Vorgabe: auch bei nicht exakter
    Schreibweise zuordnen, wenn sonst klar ist, wer gemeint ist). Passen
    mehrere unterschiedliche Betreuer auf derselben Stufe, gilt der Fall als
    mehrdeutig - dann wird NICHT auf eine lockerere Stufe weiter versucht
    (sonst waechst das Verwechslungsrisiko), sondern None geliefert (siehe
    Kunde in Manuelle_Pruefung/unmatched-Uebersicht).

    Stufen: 1) Nachname exakt + Vorname exakt, 2) Nachname exakt + Vorname
    unscharf, 3) Nachname unscharf + Vorname exakt, 4) Nachname unscharf +
    Vorname unscharf."""
    if not kunde:
        return None
    kunde = str(kunde).strip()

    candidate_pairs = []
    if "," in kunde:
        a, b = (p.strip() for p in kunde.split(",", 1))
        candidate_pairs.append((normalize_for_match(a), normalize_for_match(b)))  # Nachname, Vorname
    tokens = kunde.replace(",", " ").split()
    if len(tokens) >= 2:
        first = normalize_for_match(tokens[0])
        rest = normalize_for_match(" ".join(tokens[1:]))
        candidate_pairs.append((rest, first))  # Nachname=rest, Vorname=erstes Token
        candidate_pairs.append((first, rest))  # falls Vorname zuerst steht

    for nachname, vorname in candidate_pairs:
        if not nachname or not vorname:
            continue

        exact = [bet for vn, bet in lookup["by_surname"].get(nachname, []) if vn == vorname]
        distinct = set(exact)
        if len(distinct) == 1:
            return distinct.pop()
        if len(distinct) > 1:
            continue  # mehrdeutig auf dieser Stufe -> nicht lockerer weiterversuchen

        fuzzy_vor = [bet for vn, bet in lookup["by_surname"].get(nachname, []) if _names_match(vn, vorname)]
        distinct = set(fuzzy_vor)
        if len(distinct) == 1:
            return distinct.pop()
        if len(distinct) > 1:
            continue

        exact_vor_fuzzy_nach = [bet for nn, vn, bet in lookup["people"]
                                 if vn == vorname and nn != nachname and _names_match(nn, nachname)]
        distinct = set(exact_vor_fuzzy_nach)
        if len(distinct) == 1:
            return distinct.pop()
        if len(distinct) > 1:
            continue

        fuzzy_both = [bet for nn, vn, bet in lookup["people"]
                      if nn != nachname and _names_match(nn, nachname) and _names_match(vn, vorname)]
        distinct = set(fuzzy_both)
        if len(distinct) == 1:
            return distinct.pop()

    company_key = _strip_legal_form(normalize_for_match(kunde))
    if company_key in lookup["companies"]:
        return lookup["companies"][company_key]
    return None


def extract_bank_credits(path, password=""):
    """Liest einen VR-Bank-Rhein-Neckar-Kontoauszug (PDF mit normalem
    Textlayer) und liefert alle Buchungen (nicht nur Gutschriften) als Liste
    von Dicts: datum, vorgang, betrag (vorzeichenbehaftet), sender,
    verwendungszweck.

    Die Seiten wiederholen Brief-Kopf/Fusszeilen und einen 'Uebertrag auf/
    von Blatt N'-Saldo-Hinweis auf jeder Seite - werden vorab herausgefiltert
    und alle Seiten zu einem durchgehenden Zeilenstrom zusammengefuegt, damit
    eine Buchung, deren Absenderzeile durch einen Seitenumbruch von ihrer
    Buchungszeile getrennt ist, nicht faelschlich der naechsten Buchung oder
    einer Kopfzeile zugeordnet wird."""
    all_lines = []
    with pdfplumber.open(path, password=password) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line = line.strip()
                if not line:
                    continue
                if BANK_BOILERPLATE_RE.search(line) or BANK_BOILERPLATE_WHOLE_LINE_RE.match(line):
                    continue
                all_lines.append(line)

    txs = []
    i, n = 0, len(all_lines)
    while i < n:
        m = BANK_TX_START_RE.match(all_lines[i])
        if not m:
            i += 1
            continue
        bu_tag, _wert, vorgang, amt_s, flag = m.groups()
        amt = parse_amount(amt_s)
        amt = -amt if flag == "S" else amt
        sender = all_lines[i + 1] if i + 1 < n else ""
        j = i + 2
        detail_lines = []
        while j < n and not BANK_TX_START_RE.match(all_lines[j]):
            detail_lines.append(all_lines[j])
            j += 1
        txs.append({
            "datum": bu_tag, "vorgang": vorgang.strip(), "betrag": round(amt, 2),
            "sender": sender, "verwendungszweck": " ".join(detail_lines),
        })
        i = j
    return txs


def reconcile_bank_credits(bank_credits, insurer_keys):
    """Vergleicht Gutschriften aus dem Kontoauszug mit den in diesem Monat
    tatsaechlich verarbeiteten Versicherer-Dateien. Liefert eine Liste nicht
    zuordenbarer Gutschriften (Dicts wie extract_bank_credits(), zusaetzlich
    ohne 'vorgang') - das sind Zahlungseingaenge, zu denen keine passende
    Abrechnungs-PDF gefunden wurde, also ein Hinweis auf eine fehlende
    Abrechnung. Es wird zuerst nur der Absendername geprueft, erst wenn das
    nichts findet, zusaetzlich der Verwendungszweck (manche Versicherer
    zahlen ueber einen Pool-Absender wie 'Qualitypool' aus, der eigentliche
    Versicherer - z.B. 'AMEX' - steht dann nur im Verwendungszweck). Reine
    Absender-Suche zuerst zu versuchen vermeidet Fehltreffer durch
    Referenznummern im Verwendungszweck, die zufaellig wie ein anderer
    Versicherer-Name aussehen (z.B. eine VEMA-Referenznummer in einer
    Baloise-Zahlung)."""
    families_present = {insurer_bank_family(k) for k in insurer_keys}

    def try_match(text_norm):
        for family in families_present:
            aliases = INSURER_BANK_ALIASES.get(family, [family])
            if any(alias in text_norm for alias in aliases):
                return family
        return None

    unmatched = []
    for credit in bank_credits:
        if credit["vorgang"] != "GUTSCHRIFT":
            continue
        if try_match(normalize_for_match(credit["sender"])):
            continue
        if try_match(normalize_for_match(credit["sender"] + " " + credit["verwendungszweck"])):
            continue
        unmatched.append(credit)
    return unmatched


DIALOG_AMOUNT_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")


def find_dialog_total_hint(pdf):
    """Dialog: die Kundenpositionen-Tabelle selbst ist zu dicht/klein
    gedruckt (teils zusaetzlich durch Textmarker-Anmerkungen ueberdeckt) fuer
    zuverlaessige OCR - siehe OCR_UNRELIABLE_INSURERS. Die 'PG-Uebersicht'-
    Seite (aggregierte Summe je Produktgruppe, kein Kundenbezug) ist aber
    grossformatig genug, dass zumindest der Kontrollbetrag ("Summe gesamt")
    per OCR mit Schwellwert-Vorverarbeitung + Sparse-Text-Modus (--psm 11)
    einigermassen zuverlaessig lesbar ist - wird nur als Kontroll-Hinweis
    verwendet (Blatt "Manuelle_Pruefung"), nicht als belastbarer Wert."""
    if not OCR_AVAILABLE:
        return None
    for page in pdf.pages:
        if len(page.chars) > 0:
            continue  # keine gescannte Seite
        im = page.to_image(resolution=400).original.convert("L")
        im = ImageOps.autocontrast(im, cutoff=1)
        im = im.point(lambda p: 255 if p > 150 else 0)
        ocr_text = pytesseract.image_to_string(im, lang="deu", config="--psm 11")
        if "summe" not in ocr_text.lower():
            continue
        amounts = [parse_amount(a) for a in DIALOG_AMOUNT_RE.findall(ocr_text)]
        amounts = [a for a in amounts if a is not None]
        if amounts:
            return Counter(amounts).most_common(1)[0][0]
    return None


# Gescannte PDFs mit mehrspaltigem Karten-Layout: die OCR liest zwar Woerter,
# aber Namens- und Betragsspalten laufen dabei durcheinander (Brief-Kopfzeilen
# und Firmierungs-Zusaetze werden faelschlich als "Kunde" erkannt). Lieber
# ehrlich zur manuellen Pruefung markieren als falsche Zuordnungen liefern.
# Alte Leipziger ist NICHT mehr hier: die Scans sind sauber, siehe
# extract_alte_leipziger(). Barmenia ebenfalls nicht mehr: die Detailseiten
# lesen sich bei 400dpi/psm6 zuverlaessig, siehe extract_barmenia().
# Continentale ebenfalls nicht mehr: eigener Multi-Aufloesungs-Pfad mit
# Pflicht-Abgleich gegen den aufgedruckten "Neuer Saldo", siehe
# extract_continentale() - faellt bei Abweichung selbst automatisch auf
# manuelle Pruefung zurueck, muss deshalb nicht mehr hier pauschal
# ausgeschlossen werden.
OCR_UNRELIABLE_INSURERS = ["dialog"]

# Versicherer, die ihre Betragsspalte aus ihrer eigenen Soll/Haben-Sicht
# drucken statt aus Maklersicht - ein bei ihnen negativer Betrag ist fuer
# SSH eine Gutschrift und umgekehrt (Nutzer-Bestaetigung fuer Hiscox/
# Mannheimer, urspruenglich bei AIG anhand der aufgedruckten Summenzeile
# "...zu Ihren Gunsten..." entdeckt). Siehe SIGN_INVERTED_INSURERS-Nutzung
# in process_file().
SIGN_INVERTED_INSURERS = {"aig", "hiscox", "mannheimer"}


def insurer_name_from_filename(filename, month_folder):
    base = os.path.splitext(os.path.basename(filename))[0]
    base = re.sub(r"^Abrechnung-\d+-", "", base)
    for m in MONTHS_DE:
        base = re.sub(rf"-{m}-\d{{4}}$", "", base)
    base = base.rstrip("-")
    return base


# ---------------------------------------------------------------------------
# Hauptverarbeitung je Datei
# ---------------------------------------------------------------------------

# Bekannte Passwoerter fuer verschluesselte Versicherer-PDFs (vom Nutzer
# mitgeteilt). Schluessel = klein geschriebener insurer-Name (Substring-
# Match), Wert = Passwort.
PDF_PASSWORDS = {
    "itzehoer": "685195211",
}


def process_file(filepath, month_folder):
    insurer = insurer_name_from_filename(filepath, month_folder)
    filename = os.path.basename(filepath)

    password = next((p for k, p in PDF_PASSWORDS.items() if k in insurer.lower()), "")
    try:
        pdf = pdfplumber.open(filepath, password=password)
    except Exception as e:
        return {
            "insurer": insurer, "file": filename, "status": "fehler",
            "reason": f"PDF konnte nicht geoeffnet werden ({type(e).__name__}) - "
                      f"evtl. passwortgeschuetzt.",
            "rows": [], "total_hint": None,
        }

    try:
        full_text_parts = []
        for page in pdf.pages:
            t = page.extract_text() or ""
            full_text_parts.append(t)
        full_text = "\n".join(full_text_parts)
        insurer_lower = insurer.lower()

        if "allianz" in insurer_lower or "axa" in insurer_lower or "gothaer" in insurer_lower:
            # Gespiegelter Text wie bei is_mirrored_text() weiter unten,
            # aber mit eigenem, koordinatenbasiertem Parser statt manueller
            # Pruefung - siehe extract_allianz()/extract_axa()/extract_gothaer().
            if "allianz" in insurer_lower:
                rows = extract_allianz(pdf)
            elif "axa" in insurer_lower:
                rows = extract_axa(pdf)
            else:
                rows = extract_gothaer(pdf)
            total_hint = find_total_in_text(full_text[::-1])
            if not rows:
                return {
                    "insurer": insurer, "file": filename, "status": "sammelbeleg",
                    "reason": "Kein Kunden-Positionsdetail im PDF gefunden - "
                              "vermutlich reiner Sammelbeleg/Kontoauszug ohne "
                              "Einzelaufstellung.",
                    "rows": [], "total_hint": total_hint,
                }
            return {
                "insurer": insurer, "file": filename, "status": "ok",
                "reason": "", "rows": rows, "total_hint": total_hint,
            }

        if is_mirrored_text(full_text):
            reversed_text = full_text[::-1]
            total_hint = find_total_in_text(reversed_text)
            return {
                "insurer": insurer, "file": filename, "status": "sonderformat",
                "reason": "PDF-Text ist pro Seite gespiegelt/verdreht (bekannter "
                          "Sonderfall bei diesem Versicherer-Layout). "
                          "Automatische Kundenzuordnung in dieser Version nicht "
                          "moeglich - bitte manuell pruefen.",
                "rows": [], "total_hint": total_hint,
            }

        if any(k in insurer.lower() for k in OCR_UNRELIABLE_INSURERS):
            total_hint = find_total_in_text(full_text)
            if "dialog" in insurer.lower():
                total_hint = find_dialog_total_hint(pdf) or total_hint
            return {
                "insurer": insurer, "file": filename, "status": "sonderformat",
                "reason": "Gescanntes PDF - enthaelt echte Kundenpositionen, "
                          "aber die Texterkennung (OCR) ist zu unzuverlaessig "
                          "(vertauscht Namen/Betraege bzw. liefert zu viel "
                          "Bildrauschen). Automatische Kundenzuordnung in "
                          "dieser Version nicht moeglich - bitte manuell "
                          "pruefen, NICHT als Sammelbeleg ohne Details "
                          "missverstehen.",
                "rows": [], "total_hint": total_hint,
            }

        if insurer_lower == "continentale" and OCR_AVAILABLE:
            # Continentale 'Provisionsnote Einzelergebnisse': sehr dichte,
            # OCR-mehrdeutige Tabelle - siehe extract_continentale(). Das
            # Ergebnis wird zwingend gegen den aufgedruckten "Neuer Saldo"
            # geprueft; bei Abweichung faellt die Datei automatisch auf
            # manuelle Pruefung zurueck statt falsche Zahlen zu liefern.
            rows, target_total = extract_continentale(pdf)
            extracted_sum = round(sum(r[2] for r in rows), 2)
            if not rows or target_total is None or abs(extracted_sum - target_total) > 0.01:
                return {
                    "insurer": insurer, "file": filename, "status": "sonderformat",
                    "reason": "OCR-Ergebnis der Kundenpositionen stimmt nicht "
                              "(oder konnte nicht geprueft werden) mit dem im "
                              "PDF aufgedruckten 'Neuer Saldo' ueberein "
                              f"(extrahiert: {extracted_sum}, Soll: {target_total}) "
                              "- bitte manuell pruefen, um keine falschen "
                              "Kundenzuordnungen zu riskieren.",
                    "rows": [], "total_hint": target_total,
                }
            return {
                "insurer": insurer, "file": filename, "status": "ok",
                "reason": "OCR verwendet, gegen 'Neuer Saldo' verifiziert",
                "rows": rows, "total_hint": target_total,
            }

        if insurer_lower == "barmenia" and OCR_AVAILABLE:
            # Barmenia/Gothaer 'Vergluetungsnachweis': gescanntes PDF, dessen
            # Detailtabelle mit dem Standard-OCR-Pfad (300dpi, wortpositions-
            # basiert) die Betragsspalten verliert - eigener Pfad mit
            # hoeherer Aufloesung, siehe extract_barmenia().
            rows = extract_barmenia(pdf)
            total_hint = find_total_in_text(full_text)
            if not rows:
                return {
                    "insurer": insurer, "file": filename, "status": "sammelbeleg",
                    "reason": "Kein Kunden-Positionsdetail im PDF gefunden - "
                              "vermutlich reiner Sammelbeleg/Kontoauszug ohne "
                              "Einzelaufstellung.",
                    "rows": [], "total_hint": total_hint,
                }
            return {
                "insurer": insurer, "file": filename, "status": "ok",
                "reason": "OCR verwendet", "rows": rows, "total_hint": total_hint,
            }

        if insurer_lower == "arag" and OCR_AVAILABLE:
            # ARAG-Einzelaufstellung: gescanntes PDF, dessen Tabelle mit dem
            # Standard-OCR-Pfad (300dpi, wortpositionsbasiert) zu schlecht
            # erkannt wird - eigener Pfad mit hoeherer Aufloesung, siehe
            # extract_arag().
            rows = extract_arag(pdf)
            total_hint = find_total_in_text(full_text)
            if not rows:
                return {
                    "insurer": insurer, "file": filename, "status": "sammelbeleg",
                    "reason": "Kein Kunden-Positionsdetail im PDF gefunden - "
                              "vermutlich reiner Sammelbeleg/Kontoauszug ohne "
                              "Einzelaufstellung.",
                    "rows": [], "total_hint": total_hint,
                }
            return {
                "insurer": insurer, "file": filename, "status": "ok",
                "reason": "OCR verwendet", "rows": rows, "total_hint": total_hint,
            }

        target_month_num = None
        for m_idx, m_name in enumerate(MONTHS_DE, start=1):
            if month_folder.lower().startswith(m_name.lower()):
                target_month_num = m_idx
                break

        pages_words = []
        used_ocr = False
        for pidx, page in enumerate(pdf.pages):
            page_full_text = page.extract_text() or ""
            page_text_start = page_full_text[:60].lower()
            if "enstandsliste" in page_text_start or "vm-konto" in page_text_start:
                # "Außenstandsliste" (z.B. Mannheimer): Liste noch nicht
                # bezahlter/offener Posten, noch keine realisierte Courtage.
                # "VM-Konto" (z.B. Itzehoer): Kontobewegungs-/Saldo-Uebersicht
                # in einem anderen, oft zeichenweise gespreizten Layout ohne
                # eigene Kopfzeilen-Erkennung - wuerde sonst mit der zuletzt
                # erkannten Spaltenaufteilung Muell als Buchungszeilen liefern.
                pages_words.append((pidx, [], "text"))
                continue

            words, source = get_words_for_page(page)
            if source == "ocr":
                used_ocr = True

            if target_month_num is not None:
                # Manche PDFs, die im Monatsordner X liegen, enthalten eine
                # eingebettete Teilabrechnung fuer einen ANDEREN Monat (z.B.
                # SV Sparkassenversicherung: der Juni-PDF-Ordner enthaelt auch
                # einen "vorlaeufige Buchungen"-Abschnitt fuer 01.07.-05.07.).
                # Anhand von "Vom: DD.MM.YYYY" erkennen und ausschliessen,
                # damit diese Betraege nicht dem falschen Monat zugerechnet
                # werden. Bei gescannten Seiten (kein Text-Layer) aus den
                # ersten OCR-Woertern rekonstruieren.
                period_text = page_full_text[:300]
                if not period_text.strip():
                    period_text = " ".join(w["text"] for w in words[:100])
                m = re.search(r"Vom:?\s*(\d{2})\.(\d{2})\.(\d{4})", period_text)
                if m and int(m.group(2)) != target_month_num:
                    pages_words.append((pidx, [], "text"))
                    continue

            pages_words.append((pidx, words, source))

        if insurer_lower in SIGN_INVERTED_INSURERS:
            # AIG, Hiscox und Mannheimer drucken ihre Betragsspalte aus Sicht
            # des Versicherers (dessen eigenes Soll/Haben), nicht aus Sicht
            # von SSH als Makler - ein beim Versicherer negativer Saldo ist
            # fuer uns eine Gutschrift und umgekehrt (Nutzer-Bestaetigung).
            # Bei AIG zusaetzlich explizit belegt durch die Summenzeile
            # "Summe Vermittler (Provision/Courtage zu Ihren Gunsten): ...
            # 112,39-". Vorzeichen daher fuer alle drei explizit umdrehen.
            rows = [(p, k, -amt, l, s) for (p, k, amt, l, s) in extract_generic_table(pages_words)]
        elif "alte-leipziger" in insurer_lower:
            rows = extract_alte_leipziger(pages_words)
        elif "fondsfinanz" in insurer_lower:
            rows = extract_fondsfinanz(pages_words)
        elif "vhv" in insurer_lower:
            rows = extract_vhv(pages_words)
        elif "swiss" in insurer_lower:
            # Swiss Life liefert zwei Formate im selben PDF: eine normale
            # Tabelle (Bestandspflege) und ein Block-Format
            # (Einzelaufstellung, "VN/VP ... Summe Vertrag").
            rows = extract_generic_table(pages_words) + extract_vn_summevertrag_block(pages_words)
            vsv = extract_swiss_life_vsv_deduction(list(enumerate(full_text_parts)))
            if vsv:
                vsv_page, vsv_amount, vsv_line = vsv
                rows.append((
                    vsv_page,
                    "Vertrauensschadenversicherung-Beitrag (nicht kundenspezifisch, siehe Abrechnungsuebersicht)",
                    round(vsv_amount, 2), vsv_line, "text",
                ))
        elif "sparkassenversicherung" in insurer_lower:
            rows = extract_sv_sparkasse(pages_words)
        else:
            rows = extract_generic_table(pages_words)

        total_hint = find_total_in_text(full_text)
        if total_hint is not None and insurer_lower in SIGN_INVERTED_INSURERS:
            total_hint = -total_hint
        if not full_text.strip() and used_ocr:
            ocr_text_parts = []
            for pidx, words, source in pages_words:
                ocr_text_parts.append(" ".join(w["text"] for w in words))
            total_hint = find_total_in_text("\n".join(ocr_text_parts)) or total_hint

        if not rows:
            return {
                "insurer": insurer, "file": filename, "status": "sammelbeleg",
                "reason": "Kein Kunden-Positionsdetail im PDF gefunden - "
                          "vermutlich reiner Sammelbeleg/Kontoauszug ohne "
                          "Einzelaufstellung.",
                "rows": [], "total_hint": total_hint,
            }

        return {
            "insurer": insurer, "file": filename, "status": "ok",
            "reason": "OCR verwendet" if used_ocr else "",
            "rows": rows, "total_hint": total_hint,
        }
    finally:
        pdf.close()


def process_csv_file(filepath):
    """Verarbeitet eine VEMA-Pool-CSV (siehe extract_vema_csv). Liefert
    dasselbe Ergebnis-Dict-Format wie process_file(), damit process_files()
    PDFs und CSVs einheitlich behandeln kann."""
    filename = os.path.basename(filepath)
    rows = extract_vema_csv(filepath)
    if not rows:
        return {
            "insurer": "VEMA-Pool", "file": filename, "status": "sammelbeleg",
            "reason": "Keine erkennbaren Buchungszeilen in der CSV-Datei "
                      "gefunden (falsches Spaltenformat?).",
            "rows": [], "total_hint": None,
        }
    return {
        "insurer": "VEMA-Pool", "file": filename, "status": "ok",
        "reason": "", "rows": rows, "total_hint": None,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def process_files(files, month_label, progress_callback=None):
    """Verarbeitet eine Liste von PDF- und/oder CSV-Pfaden und liefert die
    vier Ergebnis-Tabellen (df_rows, df_control, df_agg, df_problem) als
    DataFrames. CSV-Dateien (z.B. VEMA-Poolabrechnung-N.csv) werden ueber
    process_csv_file() statt process_file() verarbeitet, siehe
    extract_vema_csv().

    progress_callback (optional): wird nach jeder Datei mit
    (index, gesamt, dateiname) aufgerufen - fuer Fortschrittsanzeigen in
    einer Oberflaeche (CLI-print oder z.B. Streamlit-Progressbar)."""
    all_rows = []
    aggregate_rows = []
    problem_rows = []
    control_rows = []

    for i, f in enumerate(files):
        if progress_callback:
            progress_callback(i, len(files), os.path.basename(f))
        if f.lower().endswith(".csv"):
            result = process_csv_file(f)
        else:
            result = process_file(f, month_label)
        insurer = result["insurer"]
        filename = result["file"]
        status = result["status"]

        if status == "ok":
            for page_idx, kunde, provision, raw_line, source in result["rows"]:
                all_rows.append({
                    "Versicherer": insurer,
                    "Kunde": kunde,
                    "Provision": provision,
                    "Datei": filename,
                    "Seite": page_idx + 1,
                    "Quelle": {"ocr": "OCR", "csv": "CSV"}.get(source, "Text"),
                    "Rohzeile": raw_line,
                })
            extracted_sum = sum(r[2] for r in result["rows"])
            control_rows.append({
                "Versicherer": insurer, "Datei": filename,
                "Anzahl_Buchungen": len(result["rows"]),
                "Summe_extrahiert": round(extracted_sum, 2),
                "Summe_lt_PDF": result["total_hint"],
                "Differenz": (round(extracted_sum - result["total_hint"], 2)
                              if result["total_hint"] is not None else None),
                "Hinweis": result["reason"],
            })
        elif status == "sammelbeleg":
            aggregate_rows.append({
                "Versicherer": insurer, "Datei": filename,
                "Betrag_lt_PDF": result["total_hint"],
                "Hinweis": result["reason"],
            })
        elif status == "sonderformat":
            problem_rows.append({
                "Versicherer": insurer, "Datei": filename,
                "Betrag_lt_PDF": result["total_hint"],
                "Grund": result["reason"],
            })
        else:
            problem_rows.append({
                "Versicherer": insurer, "Datei": filename,
                "Betrag_lt_PDF": None,
                "Grund": result["reason"],
            })

    df_rows = pd.DataFrame(all_rows)
    df_agg = pd.DataFrame(aggregate_rows)
    df_problem = pd.DataFrame(problem_rows)
    df_control = pd.DataFrame(control_rows)
    return df_rows, df_control, df_agg, df_problem


def apply_betreuer(df_rows, lookup):
    """Ergaenzt df_rows um eine Spalte 'Betreuer' (siehe match_betreuer(),
    MANUAL_BETREUER_OVERRIDES_RAW) sowie je eine Spalte pro Partner
    (PARTNER_NAMES) mit dem jeweiligen Betrag der Zeile (sonst leer) - fuer
    die farbige Kunde_Provision-Ansicht bzw. die PDF-Uebersicht. Optionaler
    Schritt: wird nur aufgerufen, wenn eine Betreuer.xlsx verfuegbar ist
    (siehe main()/app.py) - ohne diesen Aufruf bleibt df_rows unveraendert."""
    if df_rows is None or df_rows.empty:
        return df_rows
    df_rows = df_rows.copy()

    overrides = {normalize_for_match(k): v for k, v in MANUAL_BETREUER_OVERRIDES_RAW.items()}

    def resolve(kunde):
        override = overrides.get(normalize_for_match(kunde))
        return override if override else match_betreuer(kunde, lookup)

    df_rows["Betreuer"] = df_rows["Kunde"].apply(resolve)

    # Nicht-kundenspezifische Sammelabzuege (Stornoreserve, VSV-Beitrag,
    # siehe COLLECTIVE_DEDUCTION_MARKER) koennen nicht ueber einen Namen
    # zugeordnet werden - sie werden stattdessen automatisch dem Betreuer
    # der betragsmaessig groessten (bereits zugeordneten) Kundenposition
    # derselben Abrechnungsdatei zugerechnet (Nutzer-Beispiel: die
    # Fondsfinanz-Stornoreserve gehoert zu Robin Heckmann, weil dessen
    # Kunde Heckmann, Oliver die mit Abstand groesste Position in
    # derselben Abrechnung ist).
    collective_mask = df_rows["Kunde"].str.contains(COLLECTIVE_DEDUCTION_MARKER, na=False)
    for idx in df_rows[collective_mask].index:
        datei = df_rows.at[idx, "Datei"]
        same_file = df_rows[(df_rows["Datei"] == datei) & ~collective_mask & df_rows["Betreuer"].notna()]
        if not same_file.empty:
            top_row = same_file.loc[same_file["Provision"].abs().idxmax()]
            df_rows.at[idx, "Betreuer"] = top_row["Betreuer"]

    for partner in PARTNER_NAMES:
        df_rows[partner] = df_rows.apply(
            lambda r, p=partner: r["Provision"] if r["Betreuer"] == p else None, axis=1
        )
    return df_rows


def write_excel(df_rows, df_control, df_agg, df_problem, out_target, df_bank_unmatched=None):
    """Schreibt die Ergebnis-Tabellen als (optisch aufbereitete) Excel-Datei
    mit den vier Standard-Blaettern (siehe README: Aufbau der Ausgabe-Excel).
    out_target ist entweder ein Dateipfad oder ein Datei-artiges Objekt
    (z.B. io.BytesIO fuer die Web-Oberflaeche). df_bank_unmatched (optional):
    Ergebnis von reconcile_bank_credits() als DataFrame - wird als fuenftes
    Blatt "Fehlende_Abrechnungen" ergaenzt, falls angegeben."""
    if isinstance(out_target, (str, os.PathLike)):
        out_dir = os.path.dirname(out_target)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

    has_betreuer = df_rows is not None and "Betreuer" in df_rows.columns
    kunde_cols = ["Versicherer", "Kunde", "Provision", "Datei", "Seite", "Quelle", "Rohzeile"]
    kunde_currency = {"Provision"}
    if has_betreuer:
        kunde_cols = kunde_cols + ["Betreuer"] + PARTNER_NAMES
        kunde_currency = kunde_currency | set(PARTNER_NAMES)

    sheets = [
        ("Kunde_Provision", df_rows, kunde_cols, kunde_currency),
        ("Kontrolle", df_control,
         ["Versicherer", "Datei", "Anzahl_Buchungen", "Summe_extrahiert",
          "Summe_lt_PDF", "Differenz", "Hinweis"],
         {"Summe_extrahiert", "Summe_lt_PDF", "Differenz"}),
        ("Sammelbelege_ohne_Details", df_agg,
         ["Versicherer", "Datei", "Betrag_lt_PDF", "Hinweis"],
         {"Betrag_lt_PDF"}),
        ("Manuelle_Pruefung", df_problem,
         ["Versicherer", "Datei", "Betrag_lt_PDF", "Grund"],
         {"Betrag_lt_PDF"}),
    ]
    if df_bank_unmatched is not None:
        sheets.append((
            "Fehlende_Abrechnungen", df_bank_unmatched,
            ["datum", "betrag", "sender", "verwendungszweck"],
            {"betrag"},
        ))
    if has_betreuer:
        df_unmatched = df_rows[df_rows["Betreuer"].isna()][
            ["Versicherer", "Kunde", "Provision", "Datei"]
        ].drop_duplicates()
        sheets.append((
            "Kunde_ohne_Betreuer", df_unmatched,
            ["Versicherer", "Kunde", "Provision", "Datei"],
            {"Provision"},
        ))

    df_insurer_sum, df_partner_sum = _build_summary_frames(df_rows)

    with pd.ExcelWriter(out_target, engine="openpyxl") as writer:
        df_insurer_sum.to_excel(writer, sheet_name="Zusammenfassung", index=False, startrow=0)
        _style_worksheet(writer.sheets["Zusammenfassung"], df_insurer_sum,
                          ["Versicherer", "Summe"], {"Summe"})
        if df_partner_sum is not None:
            _write_partner_summary(writer.sheets["Zusammenfassung"], df_insurer_sum, df_partner_sum)

        for sheet_name, df, default_cols, currency_cols in sheets:
            (df if not df.empty else pd.DataFrame(columns=default_cols)).to_excel(
                writer, sheet_name=sheet_name, index=False, columns=default_cols
            )
            _style_worksheet(writer.sheets[sheet_name], df, default_cols, currency_cols)
            if sheet_name == "Kunde_Provision" and has_betreuer:
                _color_betreuer_rows(writer.sheets[sheet_name], df, default_cols)


def _build_summary_frames(df_rows):
    """Baut die beiden Tabellen fuer das Blatt 'Zusammenfassung': Summe je
    Versicherer (immer) sowie Summe je Partner inkl. 'Nicht zugeordnet'
    (nur wenn eine Betreuer-Zuordnung vorliegt, sonst None)."""
    if df_rows is None or df_rows.empty:
        return pd.DataFrame(columns=["Versicherer", "Summe"]), None

    df_insurer_sum = (
        df_rows.groupby("Versicherer", as_index=False)["Provision"].sum()
        .rename(columns={"Provision": "Summe"})
        .sort_values("Summe", ascending=False)
    )

    df_partner_sum = None
    if "Betreuer" in df_rows.columns:
        rows = [{"Partner": p, "Summe": df_rows.loc[df_rows["Betreuer"] == p, "Provision"].sum()}
                for p in PARTNER_NAMES]
        rows.append({"Partner": "Nicht zugeordnet",
                      "Summe": df_rows.loc[df_rows["Betreuer"].isna(), "Provision"].sum()})
        df_partner_sum = pd.DataFrame(rows)

    return df_insurer_sum, df_partner_sum


def _write_partner_summary(ws, df_insurer_sum, df_partner_sum):
    """Schreibt die 'Summe je Partner'-Tabelle unterhalb der Versicherer-
    Summen in dasselbe Blatt (eigene Kopfzeile/Formatierung, da
    _style_worksheet() von genau einer Tabelle ab Zeile 1 ausgeht)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    start_row = len(df_insurer_sum) + 3
    ws.cell(row=start_row, column=1, value="Partner")
    ws.cell(row=start_row, column=2, value="Summe")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in (1, 2):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    fills = {p: PatternFill(start_color=c, end_color=c, fill_type="solid")
             for p, c in PARTNER_COLORS_XLSX.items()}
    for i, row in enumerate(df_partner_sum.itertuples(), start=1):
        row_idx = start_row + i
        ws.cell(row=row_idx, column=1, value=row.Partner)
        cell = ws.cell(row=row_idx, column=2, value=row.Summe)
        cell.number_format = '#,##0.00 "€"'
        fill = fills.get(row.Partner)
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill
            ws.cell(row=row_idx, column=2).fill = fill

    ws.column_dimensions[get_column_letter(1)].width = max(ws.column_dimensions[get_column_letter(1)].width or 10, 20)
    ws.column_dimensions[get_column_letter(2)].width = max(ws.column_dimensions[get_column_letter(2)].width or 10, 16)


def _color_betreuer_rows(ws, df, columns):
    """Faerbt jede Zeile der Kunde_Provision-Tabelle je nach zugeordnetem
    Betreuer ein (siehe PARTNER_COLORS_XLSX, Nutzer-Vorgabe Rot/Gelb/Blau) -
    so ist beim Ueberfliegen der Tabelle auf einen Blick erkennbar, wer
    fuer welchen Umsatz verantwortlich ist. Zeilen ohne Zuordnung
    (Betreuer leer) bleiben ungefaerbt (siehe Kunde_ohne_Betreuer-Blatt)."""
    from openpyxl.styles import PatternFill

    if df is None or df.empty or "Betreuer" not in df.columns:
        return
    fills = {p: PatternFill(start_color=c, end_color=c, fill_type="solid")
             for p, c in PARTNER_COLORS_XLSX.items()}
    n_cols = len(columns)
    for row_idx, betreuer in enumerate(df["Betreuer"], start=2):
        fill = fills.get(betreuer)
        if not fill:
            continue
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _style_worksheet(ws, df, columns, currency_cols):
    """Optische Aufbereitung eines Ergebnis-Blatts: fette, farbige
    Kopfzeile, fixierte erste Zeile, Auto-Filter, sinnvolle Spaltenbreiten,
    Waehrungsformat fuer Betragsspalten sowie eine rote Hervorhebung
    abweichender "Differenz"-Werte im Kontrolle-Blatt."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    n_rows = len(df) if df is not None and not df.empty else 0
    n_cols = len(columns)
    if n_cols == 0:
        return

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"

    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        if n_rows > 0 and col_name in df.columns:
            longest_value = df[col_name].map(lambda v: len(str(v))).max()
        else:
            longest_value = 0
        width = max(len(str(col_name)), longest_value) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), 60)

        if col_name in currency_cols:
            for row_idx in range(2, n_rows + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00 "€"'

    if n_rows > 0 and "Differenz" in columns:
        diff_letter = get_column_letter(columns.index("Differenz") + 1)
        rng = f"{diff_letter}2:{diff_letter}{n_rows + 1}"
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="notEqual", formula=["0"], fill=red_fill),
        )


def _fmt_eur(x):
    """Deutsches Zahlenformat (Komma als Dezimaltrennzeichen, Punkt als
    Tausendertrennzeichen), z.B. 1234.5 -> '1.234,50 EUR'. 'EUR' statt '€'
    ausgeschrieben, da die Standard-PDF-Schrift (helvetica) das
    Euro-Zeichen nicht unterstuetzt."""
    return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " EUR"


def build_summary_pdf(df_rows, out_target, month_label):
    """Erstellt eine einseitige (bzw. mehrseitige) PDF-Uebersicht: je
    Versicherer absteigend nach Gesamtumsatz sortiert, mit der jeweiligen
    Kunden-/Vertragsaufstellung (ebenfalls absteigend nach Betrag) direkt
    daneben. Rein informativ/optisch - keine neue Extraktionslogik, nutzt
    nur die bereits extrahierten Zeilen aus process_files().

    out_target: Dateipfad (str) oder datei-artiges Objekt (z.B. io.BytesIO
    fuer die Web-Oberflaeche)."""
    from fpdf import FPDF

    pdf = FPDF(format="A4", unit="mm")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("helvetica", "B", 16)
    pdf.cell(0, 10, f"Courtage-Uebersicht {month_label}", new_x="LMARGIN", new_y="NEXT")

    if df_rows is None or df_rows.empty:
        pdf.set_font("helvetica", "", 11)
        pdf.cell(0, 8, "Keine Buchungszeilen vorhanden.")
        pdf.output(out_target)
        return

    has_betreuer = "Betreuer" in df_rows.columns
    if has_betreuer:
        per_customer = df_rows.groupby(["Versicherer", "Kunde", "Betreuer"], as_index=False,
                                        dropna=False)["Provision"].sum()
    else:
        per_customer = df_rows.groupby(["Versicherer", "Kunde"], as_index=False)["Provision"].sum()
        per_customer["Betreuer"] = None
    insurer_totals = per_customer.groupby("Versicherer")["Provision"].sum().sort_values(ascending=False)

    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 7, f"Gesamtsumme: {_fmt_eur(insurer_totals.sum())}", new_x="LMARGIN", new_y="NEXT")

    if has_betreuer:
        pdf.set_font("helvetica", "", 9)
        y_legend = pdf.get_y() + 1
        x_legend = pdf.l_margin
        for partner in PARTNER_NAMES:
            r, g, b = PARTNER_SWATCH_COLORS_PDF[partner]
            pdf.set_fill_color(r, g, b)
            pdf.rect(x_legend, y_legend + 0.5, 4, 4, "F")
            pdf.set_xy(x_legend + 5, y_legend)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(55, 5, partner)
            x_legend += 60
        pdf.set_y(y_legend + 6)

    pdf.ln(3)

    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 7, "Summe je Versicherer", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9)
    for versicherer, total in insurer_totals.items():
        pdf.cell(100, 5, versicherer)
        pdf.cell(0, 5, _fmt_eur(total), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    if has_betreuer:
        pdf.set_font("helvetica", "B", 11)
        pdf.cell(0, 7, "Monatsuebersicht", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("helvetica", "", 10)
        for partner in PARTNER_NAMES:
            total = df_rows.loc[df_rows["Betreuer"] == partner, "Provision"].sum()
            r, g, b = PARTNER_COLORS_PDF[partner]
            pdf.set_text_color(r, g, b)
            pdf.cell(60, 6, partner)
            pdf.cell(0, 6, _fmt_eur(total), new_x="LMARGIN", new_y="NEXT")
        unmatched_total = df_rows.loc[df_rows["Betreuer"].isna(), "Provision"].sum()
        pdf.set_text_color(120, 120, 120)
        pdf.cell(60, 6, "Nicht zugeordnet")
        pdf.cell(0, 6, _fmt_eur(unmatched_total), new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

    left_w, gap, right_w = 55, 5, 125
    left_margin = pdf.l_margin

    for versicherer, total in insurer_totals.items():
        customers = per_customer[per_customer["Versicherer"] == versicherer].sort_values(
            "Provision", ascending=False
        )

        # Vor jedem Block pruefen, ob der Versicherer-Name + mindestens die
        # erste Kundenzeile noch auf die aktuelle Seite passen - sonst
        # Seitenumbruch VOR dem Block (statt mittendrin) fuer ein saubereres
        # Layout.
        if pdf.get_y() + 16 > pdf.page_break_trigger:
            pdf.add_page()

        y_start = pdf.get_y()
        page_start = pdf.page_no()
        x_left = left_margin
        x_right = left_margin + left_w + gap
        right_line_h = 5

        pdf.set_xy(x_left, y_start)
        pdf.set_font("helvetica", "B", 11)
        pdf.multi_cell(left_w, 6, versicherer)
        pdf.set_x(x_left)
        pdf.set_font("helvetica", "", 9)
        pdf.multi_cell(left_w, 5, f"Gesamt:\n{_fmt_eur(total)}")
        y_left_end, page_left_end = pdf.get_y(), pdf.page_no()

        # Rechte Spalte zeilenweise mit manueller Seitenumbruch-Kontrolle
        # ausgeben (statt multi_cell()): so bleiben linke/rechte Spalte auch
        # bei einem Umbruch mitten in einer langen Kundenliste konsistent
        # positioniert, statt dass darunterliegender Inhalt auf der falschen
        # Seite landet und grosse Luecken entstehen.
        pdf.set_xy(x_right, y_start)
        pdf.set_font("helvetica", "", 9)
        y = y_start
        for row in customers.itertuples():
            if y + right_line_h > pdf.page_break_trigger:
                pdf.add_page()
                y = pdf.t_margin
            pdf.set_xy(x_right, y)
            pdf.set_text_color(*PARTNER_COLORS_PDF.get(row.Betreuer, (0, 0, 0)))
            pdf.multi_cell(right_w, right_line_h, f"{row.Kunde} - {_fmt_eur(row.Provision)}")
            pdf.set_text_color(0, 0, 0)
            y = pdf.get_y()
        y_right_end, page_right_end = y, pdf.page_no()

        # Falls die rechte Spalte einen Seitenumbruch ausgeloest hat, ist
        # y_left_end (von der vorherigen Seite) fuer die Positionierung von
        # Trennlinie/naechstem Block irrelevant - sonst wuerde eine grosse,
        # falsch platzierte Luecke auf der neuen Seite entstehen. Es zaehlt
        # immer die Spalte, die zuletzt (auf der spaeteren Seite) endet.
        if page_right_end != page_left_end:
            y_end = y_right_end if page_right_end > page_left_end else y_left_end
        else:
            y_end = max(y_left_end, y_right_end)

        pdf.set_xy(x_left, y_end + 2)
        pdf.set_draw_color(200, 200, 200)
        pdf.line(x_left, pdf.get_y(), x_right + right_w, pdf.get_y())
        pdf.ln(3)

    pdf.output(out_target)


def main():
    if len(sys.argv) < 2:
        print("Aufruf: python courtage_extraktor.py <Monatsordner z.B. Juni-2026> [--pdf-uebersicht]")
        sys.exit(1)

    month_folder = sys.argv[1]
    make_pdf_summary = "--pdf-uebersicht" in sys.argv[2:]
    input_dir = os.path.join(BASE_DIR, month_folder)
    if not os.path.isdir(input_dir):
        print(f"Ordner nicht gefunden: {input_dir}")
        sys.exit(1)

    pdf_files = sorted(glob.glob(os.path.join(input_dir, "Abrechnung-*.pdf")))
    csv_files = sorted(glob.glob(os.path.join(input_dir, "VEMA-Poolabrechnung-*.csv")))
    files = pdf_files + csv_files
    print(f"Gefunden: {len(pdf_files)} Abrechnungs-PDFs und {len(csv_files)} "
          f"VEMA-Pool-CSV(s) in {month_folder}")

    def report_progress(i, total, filename):
        print(f"  verarbeite {filename} ...")

    df_rows, df_control, df_agg, df_problem = process_files(files, month_folder, report_progress)

    betreuer_path = os.path.join(BASE_DIR, "Betreuer.xlsx")
    if os.path.isfile(betreuer_path):
        lookup = load_betreuer_lookup(betreuer_path)
        df_rows = apply_betreuer(df_rows, lookup)
        n_unmatched = int(df_rows["Betreuer"].isna().sum()) if not df_rows.empty else 0
        print(f"Betreuer.xlsx gefunden - Kunden zugeordnet ({n_unmatched} ohne Zuordnung, "
              "siehe Blatt 'Kunde_ohne_Betreuer').")
    else:
        print("Hinweis: Betreuer.xlsx nicht gefunden - keine Betreuer-Zuordnung erstellt.")

    out_dir = os.path.join(TOOL_DIR, "output", month_folder)
    out_path = os.path.join(out_dir, f"Kunde_Provision_{month_folder}.xlsx")
    write_excel(df_rows, df_control, df_agg, df_problem, out_path)

    print()
    print(f"Fertig. {len(df_rows)} Buchungszeilen aus {len(files)} Dateien extrahiert.")
    print(f"  - {len(df_agg)} Datei(en) als Sammelbeleg ohne Kundendetail")
    print(f"  - {len(df_problem)} Datei(en) zur manuellen Pruefung")
    print(f"Ausgabe: {out_path}")

    if make_pdf_summary:
        pdf_path = os.path.join(out_dir, f"Courtage-Uebersicht_{month_folder}.pdf")
        build_summary_pdf(df_rows, pdf_path, month_folder)
        print(f"PDF-Uebersicht: {pdf_path}")


if __name__ == "__main__":
    main()
